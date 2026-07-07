import json
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load the JSON from the output file
txt_path = r"C:\Users\user\.gemini\antigravity\brain\b120cd73-e77b-45ef-a520-5f577420490e\.system_generated\steps\104\output.txt"

with open(txt_path, "r", encoding="utf-8") as f:
    content = f.read()

# Clean and extract the JSON part
# Split at the Ran Playwright code section to avoid matching brackets in JS code
parts = content.split("### Ran Playwright code")
json_part = parts[0]

match = re.search(r"(\[.*\]|\{.*\})", json_part, re.DOTALL)
if not match:
    raise ValueError("JSON block not found in the output file")

json_str = match.group(1)
data = json.loads(json_str)

# Create workbook
wb = openpyxl.Workbook()

# Colors
DARK_BLUE = "1F497D"
WHITE = "FFFFFF"
LIGHT_BLUE_LVL0 = "B8CCE4"
LIGHT_BLUE_LVL1 = "DCE6F1"
LIGHT_GRAY_LVL2 = "F2F2F2"
WHITE_COLOR = "FFFFFF"

# Fonts
font_title = Font(name="Segoe UI", size=16, bold=True, color="1F497D")
font_meta_label = Font(name="Segoe UI", size=10, bold=True)
font_meta_val = Font(name="Segoe UI", size=10)
font_header = Font(name="Segoe UI", size=10, bold=True, color=WHITE)
font_lvl0 = Font(name="Segoe UI", size=10, bold=True)
font_lvl1 = Font(name="Segoe UI", size=10, bold=True)
font_lvl2 = Font(name="Segoe UI", size=10, italic=False)
font_lvl3 = Font(name="Segoe UI", size=9, italic=True, color="595959")
font_hierarchy = Font(name="Segoe UI", size=9, color="333333")

# Fills
fill_header = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
fill_lvl0 = PatternFill(start_color=LIGHT_BLUE_LVL0, end_color=LIGHT_BLUE_LVL0, fill_type="solid")
fill_lvl1 = PatternFill(start_color=LIGHT_BLUE_LVL1, end_color=LIGHT_BLUE_LVL1, fill_type="solid")
fill_lvl2 = PatternFill(start_color=LIGHT_GRAY_LVL2, end_color=LIGHT_GRAY_LVL2, fill_type="solid")
fill_lvl3 = PatternFill(start_color=WHITE_COLOR, end_color=WHITE_COLOR, fill_type="solid")

# Alignments
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
align_hierarchy = Alignment(horizontal="left", vertical="center", wrap_text=True)

def align_with_indent(level):
    return Alignment(horizontal="left", vertical="center", indent=level * 2)

# Borders
thin_side = Side(border_style="thin", color="D3D3D3")
border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# ----------------- SHEET 1: RINGKASAN -----------------
ws_summary = wb.active
ws_summary.title = "Ringkasan Pagu"
ws_summary.views.sheetView[0].showGridLines = True

# Title
ws_summary["A1"] = "RINGKASAN PAGU ANGGARAN SATKER TAHUN 2026"
ws_summary["A1"].font = font_title
ws_summary.merge_cells("A1:J1")

# KPI Block (Total Pagu)
ws_summary["A3"] = "TOTAL PAGU SATKER"
ws_summary["A3"].font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
ws_summary["A3"].fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
ws_summary["A3"].alignment = align_center

# We will calculate total pagu during iteration
total_satker_pagu = 0

# Table Headers
headers_summary = [
    "No",
    "Kode Program",
    "Nama Program",
    "Kode Kegiatan",
    "Nama Kegiatan",
    "Kode KRO",
    "Nama KRO",
    "Kode Komponen",
    "Nama Komponen",
    "Pagu Komponen"
]

header_row = 6
for col_idx, h in enumerate(headers_summary, start=1):
    cell = ws_summary.cell(row=header_row, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_cell

# Populate summary table
row_idx = 7
num = 1
for prog in data:
    prog_code = prog["text"].split("]")[0].replace("[", "").strip()
    prog_name = prog["text"].split("]")[1].strip()
    
    for keg in prog.get("kegiatans", []):
        for out in keg.get("outputs", []):
            for komp in out.get("komponens", []):
                # Write rows
                ws_summary.cell(row=row_idx, column=1, value=num).alignment = align_center
                ws_summary.cell(row=row_idx, column=2, value=prog_code).alignment = align_center
                ws_summary.cell(row=row_idx, column=3, value=prog_name).alignment = align_left
                ws_summary.cell(row=row_idx, column=4, value=keg["code"]).alignment = align_center
                ws_summary.cell(row=row_idx, column=5, value=keg["name"]).alignment = align_left
                ws_summary.cell(row=row_idx, column=6, value=out["code"]).alignment = align_center
                ws_summary.cell(row=row_idx, column=7, value=out["name"]).alignment = align_left
                ws_summary.cell(row=row_idx, column=8, value=komp["code"]).alignment = align_center
                ws_summary.cell(row=row_idx, column=9, value=komp["name"]).alignment = align_left
                
                c_pagu = ws_summary.cell(row=row_idx, column=10, value=komp["pagu"])
                c_pagu.number_format = '#,##0'
                c_pagu.alignment = align_right
                
                # Borders and fonts
                for col_c in range(1, 11):
                    cell = ws_summary.cell(row=row_idx, column=col_c)
                    cell.border = border_cell
                    cell.font = Font(name="Segoe UI", size=10)
                
                total_satker_pagu += komp["pagu"]
                row_idx += 1
                num += 1

# Total Row
ws_summary.cell(row=row_idx, column=9, value="TOTAL").font = Font(name="Segoe UI", size=10, bold=True)
ws_summary.cell(row=row_idx, column=9).alignment = align_right
ws_summary.cell(row=row_idx, column=9).border = border_cell

c_tot = ws_summary.cell(row=row_idx, column=10, value=total_satker_pagu)
c_tot.font = Font(name="Segoe UI", size=10, bold=True)
c_tot.number_format = '#,##0'
c_tot.alignment = align_right
c_tot.border = border_cell

# Set KPI Card Value
ws_summary["A4"] = total_satker_pagu
ws_summary["A4"].font = Font(name="Segoe UI", size=18, bold=True, color="1F497D")
ws_summary["A4"].number_format = 'Rp #,##0'
ws_summary["A4"].alignment = align_center
ws_summary.merge_cells("A3:C3")
ws_summary.merge_cells("A4:C4")

# Auto widths
for col in ws_summary.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.row < 6:
            continue
        val_str = str(cell.value or '')
        if len(val_str) > max_len:
            max_len = len(val_str)
    ws_summary.column_dimensions[col_letter].width = max(max_len + 4, 10)

ws_summary.column_dimensions['C'].width = 30
ws_summary.column_dimensions['E'].width = 30
ws_summary.column_dimensions['G'].width = 30
ws_summary.column_dimensions['I'].width = 30


# ----------------- SHEET 2 & 3: DETAIL PER PROGRAM -----------------
for prog in data:
    prog_code = prog["text"].split("]")[0].replace("[", "").strip()
    sheet_name = f"Detail - {prog_code}"
    ws = wb.create_sheet(title=sheet_name)
    ws.views.sheetView[0].showGridLines = True
    
    # Title
    ws["A1"] = f"DETAIL RENCANA KERJA ANGGARAN (RKA) - PROGRAM {prog_code}"
    ws["A1"].font = font_title
    ws.merge_cells("A1:M1")
    
    ws["A3"] = "Program:"
    ws["A3"].font = font_meta_label
    ws["B3"] = prog["text"]
    ws["B3"].font = font_meta_val
    ws.merge_cells("B3:M3")
    
    # Table Headers template
    headers_det = [
        "Kegiatan",
        "KRO (Output)",
        "Komponen",
        "Kode (P/K/O/SO/K/SK/A/D)",
        "Uraian",
        "Uraian Sebelum Revisi",
        "Pagu",
        "Pagu Sebelum Revisi",
        "P",
        "S",
        "Multiyears",
        "NP",
        "Gaji"
    ]
    
    # Write Headers
    header_row = 5
    for col_idx, h in enumerate(headers_det, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_cell
        
    curr_row = 6
    
    for keg in prog.get("kegiatans", []):
        keg_text = f"[{keg['code']}] {keg['name']}"
        for out in keg.get("outputs", []):
            kro_text = f"[{out['code']}] {out['name']}"
            for komp in out.get("komponens", []):
                komp_text = f"[{komp['code']}] {komp['name']}"
                
                start_row = curr_row
                
                # Write Table rows for this Component
                for row_data in komp.get("rows", []):
                    level = row_data["level"]
                    code = row_data["code"]
                    desc = row_data["desc"]
                    desc_prev = row_data["descPrev"]
                    pagu = row_data["pagu"]
                    pagu_prev = row_data["paguPrev"]
                    
                    p_ch = row_data["p_ch"]
                    s_ch = row_data["s_ch"]
                    my_ch = row_data["my_ch"]
                    np_ch = row_data["np_ch"]
                    gj_ch = row_data["gj_ch"]
                    
                    # Columns D to M (indices 4 to 13)
                    ws.cell(row=curr_row, column=4, value=code)
                    ws.cell(row=curr_row, column=5, value=desc)
                    ws.cell(row=curr_row, column=6, value=desc_prev)
                    
                    if isinstance(pagu, (int, float)):
                        ws.cell(row=curr_row, column=7, value=pagu).number_format = '#,##0'
                    else:
                        ws.cell(row=curr_row, column=7, value=pagu)
                        
                    if isinstance(pagu_prev, (int, float)):
                        ws.cell(row=curr_row, column=8, value=pagu_prev).number_format = '#,##0'
                    else:
                        ws.cell(row=curr_row, column=8, value=pagu_prev)
                        
                    ws.cell(row=curr_row, column=9, value="✓" if p_ch else "")
                    ws.cell(row=curr_row, column=10, value="✓" if s_ch else "")
                    ws.cell(row=curr_row, column=11, value="✓" if my_ch else "")
                    ws.cell(row=curr_row, column=12, value="✓" if np_ch else "")
                    ws.cell(row=curr_row, column=13, value="✓" if gj_ch else "")
                    
                    # Formatting based on level for details columns
                    if level == 0:
                        row_fill = fill_lvl0
                        row_font = font_lvl0
                    elif level == 1:
                        row_fill = fill_lvl1
                        row_font = font_lvl1
                    elif level == 2:
                        row_fill = fill_lvl2
                        row_font = font_lvl2
                    else:
                        row_fill = fill_lvl3
                        row_font = font_lvl3
                        
                    for col_idx in range(4, 14):
                        cell = ws.cell(row=curr_row, column=col_idx)
                        cell.fill = row_fill
                        cell.font = row_font
                        cell.border = border_cell
                        
                        if col_idx == 4:
                            cell.alignment = align_center
                        elif col_idx == 5:
                            cell.alignment = align_with_indent(level)
                        elif col_idx == 6:
                            cell.alignment = align_with_indent(level)
                        elif col_idx in (7, 8):
                            cell.alignment = align_right
                        else:
                            cell.alignment = align_center
                    
                    curr_row += 1
                
                end_row = curr_row - 1
                
                # Now merge and write the hierarchy values in A, B, C (cols 1, 2, 3)
                if end_row >= start_row:
                    ws.cell(row=start_row, column=1, value=keg_text)
                    ws.cell(row=start_row, column=2, value=kro_text)
                    ws.cell(row=start_row, column=3, value=komp_text)
                    
                    # Set border and alignment for cols 1, 2, 3 across all rows in the block
                    for r in range(start_row, end_row + 1):
                        for c in (1, 2, 3):
                            cell = ws.cell(row=r, column=c)
                            cell.font = font_hierarchy
                            cell.alignment = align_hierarchy
                            cell.border = border_cell
                            # Let's set a very light background fill for the hierarchy columns to separate them
                            cell.fill = PatternFill(start_color="FAFBFD", end_color="FAFBFD", fill_type="solid")
                    
                    if end_row > start_row:
                        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                        ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
                        ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)

    # Column dimensions for details sheet
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 45
    ws.column_dimensions['F'].width = 45
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    for col_c in range(9, 14):
        col_letter = get_column_letter(col_c)
        ws.column_dimensions[col_letter].width = 8

wb.save("rekap_rka_lengkap.xlsx")
print("Consolidated Excel generated successfully!")
