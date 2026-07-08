import json
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Paths
rka_txt_path = r"C:\Users\user\.gemini\antigravity\brain\b120cd73-e77b-45ef-a520-5f577420490e\.system_generated\steps\104\output.txt"
rup_txt_path = r"C:\Users\user\.gemini\antigravity\brain\b120cd73-e77b-45ef-a520-5f577420490e\.system_generated\steps\160\output.txt"
rup_det_path = r"C:\Users\user\.gemini\antigravity\brain\b120cd73-e77b-45ef-a520-5f577420490e\.system_generated\steps\204\output.txt"

# Robust MAK parser
def parse_mak(mak):
    if not mak:
        return None
    mak = str(mak).replace(" ", "").strip()
    parts = mak.split(".")
    # If prefixed with year and satker code (9 parts)
    if len(parts) >= 9 and parts[0].isdigit() and parts[1].isdigit():
        return {
            "prog": parts[2],
            "keg": parts[3],
            "out": parts[4],
            "ro": parts[5],
            "komp": parts[6],
            "subkomp": parts[7],
            "akun": parts[8],
            "comp_key": f"{parts[2]}.{parts[3]}.{parts[4]}.{parts[5]}.{parts[6]}",
            "key": f"{parts[2]}.{parts[3]}.{parts[4]}.{parts[5]}.{parts[6]}.{parts[7]}.{parts[8]}"
        }
    # Direct MAK (7 parts)
    if len(parts) >= 7:
        return {
            "prog": parts[0],
            "keg": parts[1],
            "out": parts[2],
            "ro": parts[3],
            "komp": parts[4],
            "subkomp": parts[5],
            "akun": parts[6],
            "comp_key": f"{parts[0]}.{parts[1]}.{parts[2]}.{parts[3]}.{parts[4]}",
            "key": f"{parts[0]}.{parts[1]}.{parts[2]}.{parts[3]}.{parts[4]}.{parts[5]}.{parts[6]}"
        }
    return None

# 1. Load RKA Data
with open(rka_txt_path, "r", encoding="utf-8") as f:
    rka_content = f.read()

rka_match = re.search(r"(\[.*\]|\{.*\})", rka_content.split("### Ran Playwright code")[0], re.DOTALL)
if not rka_match:
    raise ValueError("RKA JSON not found")
rka_data = json.loads(rka_match.group(1))

# 2. Load RUP Meta Data
with open(rup_txt_path, "r", encoding="utf-8") as f:
    rup_content = f.read()

rup_match = re.search(r"(\[.*\]|\{.*\})", rup_content.split("### Ran Playwright code")[0], re.DOTALL)
if not rup_match:
    rup_match = re.search(r"(\[.*\]|\{.*\})", rup_content, re.DOTALL)

if not rup_match:
    raise ValueError("RUP JSON not found")

rup_json = json.loads(rup_match.group(1))
rup_rows = rup_json.get("rows", [])

rup_packets = []
for row in rup_rows:
    rup_packets.append({
        "id": row[0],
        "keg_name": row[1],
        "name": row[2],
        "pagu": int(row[3]) if str(row[3]).isdigit() else 0,
        "waktu": row[4],
        "sumber_dana": row[5],
        "aktif": row[6] == "true",
        "fd": row[7] == "true",
        "umumkan": row[8] == "true",
        "mak": row[17] if len(row) > 17 else ""
    })

# 3. Load RUP Granular Detailed Sub-lines
with open(rup_det_path, "r", encoding="utf-8") as f:
    det_content = f.read()

parts = det_content.split("### Ran Playwright code")
det_part = parts[0]

det_match = re.search(r"(\[.*\]|\{.*\})", det_part, re.DOTALL)
if not det_match:
    raise ValueError("Granular RUP details JSON not found")

rup_details = json.loads(det_match.group(1))

# Process RUP detailed sub-lines and index them
rup_all_lines = []
for pkt_id, items in rup_details.items():
    packet_info = next((p for p in rup_packets if p["id"] == pkt_id), None)
    packet_name = packet_info["name"] if packet_info else ""
    is_terumumkan = (packet_info["aktif"] and packet_info["fd"] and packet_info["umumkan"]) if packet_info else False
    
    for item in items:
        mak = item.get("mak", "")
        pagu = item.get("pagu", 0)
        
        parsed = parse_mak(mak)
        if parsed:
            rup_all_lines.append({
                "packet_id": pkt_id,
                "packet_name": packet_name,
                "mak": mak,
                "pagu": pagu,
                "prog": parsed["prog"],
                "keg": parsed["keg"],
                "out": parsed["out"],
                "ro": parsed["ro"],
                "komp": parsed["komp"],
                "subkomp": parsed["subkomp"],
                "akun": parsed["akun"],
                "comp_key": parsed["comp_key"],
                "key": parsed["key"],
                "is_terumumkan": is_terumumkan
            })

# Create workbook
wb = openpyxl.Workbook()

# Styles
DARK_BLUE = "1F497D"
WHITE = "FFFFFF"
LIGHT_BLUE_LVL0 = "B8CCE4"
LIGHT_BLUE_LVL1 = "DCE6F1"
LIGHT_GRAY_LVL2 = "F2F2F2"
LIGHT_GRAY_LVL3 = "F9F9F9"
LIGHT_GREEN = "E2EFDA"
LIGHT_RED = "FCE4D6"
WHITE_COLOR = "FFFFFF"
LIGHT_ORANGE = "FFF2CC"

# Fonts
font_title = Font(name="Segoe UI", size=16, bold=True, color="1F497D")
font_meta_label = Font(name="Segoe UI", size=10, bold=True)
font_meta_val = Font(name="Segoe UI", size=10)
font_header = Font(name="Segoe UI", size=10, bold=True, color=WHITE)
font_lvl0 = Font(name="Segoe UI", size=10, bold=True)
font_lvl1 = Font(name="Segoe UI", size=10, bold=True)
font_lvl2 = Font(name="Segoe UI", size=10, italic=False)
font_lvl3 = Font(name="Segoe UI", size=9, italic=True, color="595959")
font_hierarchy = Font(name="Segoe UI", size=9, color="333333")
font_match = Font(name="Segoe UI", size=9, italic=True, color="2E7D32") # green italic for matched rows
font_np = Font(name="Segoe UI", size=9, italic=True, color="7F7F7F") # gray italic for NP rows
font_draft = Font(name="Segoe UI", size=9, italic=True, color="C95D00") # Dark Orange for draft/cancelled warning

# Fills
fill_header = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
fill_lvl0 = PatternFill(start_color=LIGHT_BLUE_LVL0, end_color=LIGHT_BLUE_LVL0, fill_type="solid")
fill_lvl1 = PatternFill(start_color=LIGHT_BLUE_LVL1, end_color=LIGHT_BLUE_LVL1, fill_type="solid")
fill_lvl2 = PatternFill(start_color=LIGHT_GRAY_LVL2, end_color=LIGHT_GRAY_LVL2, fill_type="solid")
fill_lvl3 = PatternFill(start_color=WHITE_COLOR, end_color=WHITE_COLOR, fill_type="solid")
fill_sub_ro = PatternFill(start_color=LIGHT_GRAY_LVL3, end_color=LIGHT_GRAY_LVL3, fill_type="solid")

fill_green = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
fill_red = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
fill_draft = PatternFill(start_color=LIGHT_ORANGE, end_color=LIGHT_ORANGE, fill_type="solid")

# Alignments
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
align_hierarchy = Alignment(horizontal="left", vertical="center", wrap_text=True)

def align_with_indent(level):
    return Alignment(horizontal="left", vertical="center", indent=level * 2)

# Borders
thin_side = Side(border_style="thin", color="D3D3D3")
border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)


# --- 0. PRE-CALCULATE ALL TOTALS & STATS ---
total_satker_pagu = 0
total_non_pengadaan = 0
total_target_pengadaan = 0
total_rup_pagu = 0

# Sum total RKA pagus and non-pengadaan
for prog in rka_data:
    prog_code = prog["text"].split("]")[0].replace("[", "").strip()
    for keg in prog.get("kegiatans", []):
        for out in keg.get("outputs", []):
            suboutputs = out.get("suboutputs", []) or [{"id": 0, "name": out["name"], "code": "000", "komponens": out.get("komponens", [])}]
            for ro in suboutputs:
                for komp in ro.get("komponens", []):
                    total_satker_pagu += komp.get("pagu", 0)
                    
                    komp_np = False
                    komp_gj = False
                    subkomp_np = False
                    subkomp_gj = False
                    akun_np = False
                    akun_gj = False
                    
                    for r in komp.get("rows", []):
                        lvl = r.get("level", 0)
                        if lvl == 0:
                            komp_np = r.get("np_ch", False)
                            komp_gj = r.get("gj_ch", False)
                        elif lvl == 1:
                            subkomp_np = r.get("np_ch", False)
                            subkomp_gj = r.get("gj_ch", False)
                        elif lvl == 2:
                            akun_np = r.get("np_ch", False)
                            akun_gj = r.get("gj_ch", False)
                        elif lvl == 3:
                            is_np = r.get("np_ch", False) or akun_np or subkomp_np or komp_np
                            is_gj = r.get("gj_ch", False) or akun_gj or subkomp_gj or komp_gj
                            if is_np or is_gj:
                                total_non_pengadaan += r.get("pagu", 0)

total_target_pengadaan = total_satker_pagu - total_non_pengadaan

# Sum RUP totals (Only fully announced ones)
for prog in rka_data:
    prog_code = prog["text"].split("]")[0].replace("[", "").strip()
    for keg in prog.get("kegiatans", []):
        for out in keg.get("outputs", []):
            suboutputs = out.get("suboutputs", []) or [{"id": 0, "name": out["name"], "code": "000", "komponens": out.get("komponens", [])}]
            for ro in suboutputs:
                for komp in ro.get("komponens", []):
                    comp_key = f"{prog_code}.{keg['code']}.{out['code']}.{ro['code']}.{komp['code']}"
                    comp_rup = [l for l in rup_all_lines if l["comp_key"] == comp_key and l["is_terumumkan"]]
                    total_rup_pagu += sum(l["pagu"] for l in comp_rup)

# Gather RKA detailed Akun keys (7-parts)
rka_detailed_keys = set()
for prog in rka_data:
    prog_code = prog["text"].split("]")[0].replace("[", "").strip()
    for keg in prog.get("kegiatans", []):
        for out in keg.get("outputs", []):
            suboutputs = out.get("suboutputs", []) or [{"id": 0, "name": out["name"], "code": "000", "komponens": out.get("komponens", [])}]
            for ro in suboutputs:
                for komp in ro.get("komponens", []):
                    comp_key = f"{prog_code}.{keg['code']}.{out['code']}.{ro['code']}.{komp['code']}"
                    subkomp_code = ""
                    for r in komp.get("rows", []):
                        lvl = r.get("level", 0)
                        if lvl == 1:
                            subkomp_code = r["code"]
                        elif lvl == 2:
                            rka_detailed_keys.add(f"{comp_key}.{subkomp_code}.{r['code']}")

unmatched_rup_lines = []
for line in rup_all_lines:
    if line["key"] not in rka_detailed_keys:
        unmatched_rup_lines.append({"packet_id": line["packet_id"], "mak": line["mak"]})
for p in rup_packets:
    p_lines = [l for l in rup_all_lines if l["packet_id"] == p["id"]]
    if not p_lines:
        parsed = parse_mak(p["mak"])
        if parsed and parsed["key"] not in rka_detailed_keys:
            unmatched_rup_lines.append({"packet_id": p["id"], "mak": p["mak"] or "(Kosong)"})
        elif not parsed:
            unmatched_rup_lines.append({"packet_id": p["id"], "mak": p["mak"] or "(Kosong)"})

seen_unmatched = set()
unique_unmatched = []
for item in unmatched_rup_lines:
    k = f"{item['packet_id']}_{item['mak']}"
    if k not in seen_unmatched:
        seen_unmatched.add(k)
        unique_unmatched.append(item)
unique_unmatched_count = len(unique_unmatched)

# ----------------- SHEET 0: DASHBOARD EVALUASI -----------------
ws_dash = wb.active
ws_dash.title = "Dashboard Evaluasi"
ws_dash.views.sheetView[0].showGridLines = True

ws_dash["A1"] = "DASHBOARD MONITORING & EVALUASI INTEGRASI RKA-RUP"
ws_dash["A1"].font = font_title
ws_dash.merge_cells("A1:H1")

ws_dash["A2"] = "Tanggal Ekspor: 08/07/2026 | Satker: Sekretariat Badan Penyuluhan dan Pengembangan SDM Kelautan dan Perikanan"
ws_dash["A2"].font = Font(name="Segoe UI", size=10, italic=True, color="595959")
ws_dash.merge_cells("A2:H2")

# Card 1: TOTAL PAGU SATKER (A4:B6)
ws_dash.merge_cells("A4:B4")
ws_dash["A4"] = "TOTAL PAGU SATKER (A)"
ws_dash["A4"].font = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
ws_dash["A4"].fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
ws_dash["A4"].alignment = align_center

ws_dash.merge_cells("A5:B6")
ws_dash["A5"] = total_satker_pagu
ws_dash["A5"].font = Font(name="Segoe UI", size=14, bold=True, color="1F497D")
ws_dash["A5"].number_format = 'Rp #,##0'
ws_dash["A5"].alignment = align_center
for r in range(5, 7):
    for c in range(1, 3):
        ws_dash.cell(row=r, column=c).border = border_cell

# Card 2: TARGET PENGADAAN (C4:D6)
ws_dash.merge_cells("C4:D4")
ws_dash["C4"] = "TARGET PENGADAAN (C = A - B)"
ws_dash["C4"].font = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
ws_dash["C4"].fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
ws_dash["C4"].alignment = align_center

ws_dash.merge_cells("C5:D6")
ws_dash["C5"] = total_target_pengadaan
ws_dash["C5"].font = Font(name="Segoe UI", size=14, bold=True, color="203764")
ws_dash["C5"].number_format = 'Rp #,##0'
ws_dash["C5"].alignment = align_center
for r in range(5, 7):
    for c in range(3, 5):
        ws_dash.cell(row=r, column=c).border = border_cell

# Card 3: RUP TERUMUMKAN (E4:F6)
ws_dash.merge_cells("E4:F4")
ws_dash["E4"] = "RUP TERUMUMKAN (D)"
ws_dash["E4"].font = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
ws_dash["E4"].fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
ws_dash["E4"].alignment = align_center

ws_dash.merge_cells("E5:F6")
ws_dash["E5"] = total_rup_pagu
ws_dash["E5"].font = Font(name="Segoe UI", size=14, bold=True, color="375623")
ws_dash["E5"].number_format = 'Rp #,##0'
ws_dash["E5"].alignment = align_center
for r in range(5, 7):
    for c in range(5, 7):
        ws_dash.cell(row=r, column=c).border = border_cell

# Card 4: PERSENTASE CAPAIAN (G4:H6)
ws_dash.merge_cells("G4:H4")
ws_dash["G4"] = "PERSENTASE CAPAIAN (%)"
ws_dash["G4"].font = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
ws_dash["G4"].fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
ws_dash["G4"].alignment = align_center

tot_pct = total_rup_pagu / total_target_pengadaan if total_target_pengadaan > 0 else 0
ws_dash.merge_cells("G5:H6")
ws_dash["G5"] = tot_pct
ws_dash["G5"].font = Font(name="Segoe UI", size=16, bold=True, color="2E7D32" if tot_pct == 1 else "C00000")
ws_dash["G5"].number_format = '0.0%'
ws_dash["G5"].alignment = align_center
for r in range(5, 7):
    for c in range(7, 9):
        ws_dash.cell(row=r, column=c).border = border_cell

# Section Title
ws_dash["A8"] = "STATISTIK EVALUASI PAKET RKA-RUP"
ws_dash["A8"].font = Font(name="Segoe UI", size=12, bold=True, color="1F497D")

# Table Headers
stat_headers = ["No", "Indikator Evaluasi Anggaran & RUP", "Nilai / Jumlah", "Satuan"]
for col in range(1, 5):
    cell = ws_dash.cell(row=10, column=col)
    cell.value = stat_headers[col-1]
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_cell

# Stats Rows
stats_data = [
    ["1", "Total Belanja Non-Pengadaan (NP / Gaji Satker) (B)", total_non_pengadaan, "Rupiah (Rp)"],
    ["2", "Jumlah Paket RUP Terdaftar (Penyedia)", len(rup_packets), "Paket"],
    ["3", "Jumlah Paket RUP Sah (Terumumkan KPA)", len([p for p in rup_packets if p["aktif"] and p["fd"] and p["umumkan"]]), "Paket"],
    ["4", "Jumlah Paket RUP Belum Terumumkan (Draft / Dibatalkan)", len([p for p in rup_packets if not (p["aktif"] and p["fd"] and p["umumkan"])]), "Paket"],
    ["5", "Jumlah Paket RUP Kehilangan Sandingan (Potensi Salah Input MAK)", unique_unmatched_count, "Paket"]
]

for i, row in enumerate(stats_data):
    r_idx = 11 + i
    ws_dash.cell(row=r_idx, column=1, value=row[0]).alignment = align_center
    ws_dash.cell(row=r_idx, column=2, value=row[1]).alignment = align_left
    
    val_cell = ws_dash.cell(row=r_idx, column=3, value=row[2])
    if row[3] == "Rupiah (Rp)":
        val_cell.number_format = '#,##0'
        val_cell.alignment = align_right
    else:
        val_cell.number_format = '#,##0'
        val_cell.alignment = align_center
        
    ws_dash.cell(row=r_idx, column=4, value=row[3]).alignment = align_center
    
    for c in range(1, 5):
        cell = ws_dash.cell(row=r_idx, column=c)
        cell.font = Font(name="Segoe UI", size=9)
        cell.border = border_cell
        if c == 3 and (i == 3 or i == 4) and row[2] > 0:
            cell.fill = fill_draft
            cell.font = Font(name="Segoe UI", size=9, bold=True, color="C00000")

ws_dash.column_dimensions['A'].width = 5
ws_dash.column_dimensions['B'].width = 50
ws_dash.column_dimensions['C'].width = 25
ws_dash.column_dimensions['D'].width = 15
ws_dash.column_dimensions['E'].width = 18
ws_dash.column_dimensions['F'].width = 18
ws_dash.column_dimensions['G'].width = 18
ws_dash.column_dimensions['H'].width = 18

# ----------------- SHEET 1: RINGKASAN PAGU -----------------
ws_summary = wb.create_sheet(title="Ringkasan Pagu")
ws_summary.views.sheetView[0].showGridLines = True

ws_summary["A1"] = "RINGKASAN PAGU ANGGARAN SATKER TAHUN 2026"
ws_summary["A1"].font = font_title
ws_summary.merge_cells("A1:L1")

# KPI Block (Total Pagu)
ws_summary["A3"] = "TOTAL PAGU SATKER"
ws_summary["A3"].font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
ws_summary["A3"].fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
ws_summary["A3"].alignment = align_center

total_satker_pagu = 0

headers_summary = [
    "No", "Kode Program", "Nama Program", "Kode Kegiatan", "Nama Kegiatan",
    "Kode KRO", "Nama KRO", "Kode RO", "Nama RO", "Kode Komponen", "Nama Komponen", "Pagu Komponen"
]

header_row = 6
for col_idx, h in enumerate(headers_summary, start=1):
    cell = ws_summary.cell(row=header_row, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_cell

row_idx = 7
num = 1
for prog in rka_data:
    prog_code = prog["text"].split("]")[0].replace("[", "").strip()
    prog_name = prog["text"].split("]")[1].strip()
    
    for keg in prog.get("kegiatans", []):
        for out in keg.get("outputs", []):
            suboutputs_list = out.get("suboutputs", [])
            if not suboutputs_list:
                suboutputs_list = [{ "id": 0, "name": out["name"], "code": "000", "komponens": out.get("komponens", []) }]
                
            for ro in suboutputs_list:
                for komp in ro.get("komponens", []):
                    ws_summary.cell(row=row_idx, column=1, value=num).alignment = align_center
                    ws_summary.cell(row=row_idx, column=2, value=prog_code).alignment = align_center
                    ws_summary.cell(row=row_idx, column=3, value=prog_name).alignment = align_left
                    ws_summary.cell(row=row_idx, column=4, value=keg["code"]).alignment = align_center
                    ws_summary.cell(row=row_idx, column=5, value=keg["name"]).alignment = align_left
                    ws_summary.cell(row=row_idx, column=6, value=out["code"]).alignment = align_center
                    ws_summary.cell(row=row_idx, column=7, value=out["name"]).alignment = align_left
                    ws_summary.cell(row=row_idx, column=8, value=ro["code"]).alignment = align_center
                    ws_summary.cell(row=row_idx, column=9, value=ro["name"]).alignment = align_left
                    ws_summary.cell(row=row_idx, column=10, value=komp["code"]).alignment = align_center
                    ws_summary.cell(row=row_idx, column=11, value=komp["name"]).alignment = align_left
                    
                    c_pagu = ws_summary.cell(row=row_idx, column=12, value=komp["pagu"])
                    c_pagu.number_format = '#,##0'
                    c_pagu.alignment = align_right
                    
                    for col_c in range(1, 13):
                        cell = ws_summary.cell(row=row_idx, column=col_c)
                        cell.border = border_cell
                        cell.font = Font(name="Segoe UI", size=9)
                    
                    total_satker_pagu += komp["pagu"]
                    row_idx += 1
                    num += 1

ws_summary.cell(row=row_idx, column=11, value="TOTAL").font = Font(name="Segoe UI", size=10, bold=True)
ws_summary.cell(row=row_idx, column=11).alignment = align_right
ws_summary.cell(row=row_idx, column=11).border = border_cell

c_tot = ws_summary.cell(row=row_idx, column=12, value=total_satker_pagu)
c_tot.font = Font(name="Segoe UI", size=10, bold=True)
c_tot.number_format = '#,##0'
c_tot.alignment = align_right
c_tot.border = border_cell

ws_summary["A4"] = total_satker_pagu
ws_summary["A4"].font = Font(name="Segoe UI", size=18, bold=True, color="1F497D")
ws_summary["A4"].number_format = 'Rp #,##0'
ws_summary["A4"].alignment = align_center
ws_summary.merge_cells("A3:C3")
ws_summary.merge_cells("A4:C4")

ws_summary.column_dimensions['A'].width = 5
ws_summary.column_dimensions['B'].width = 12
ws_summary.column_dimensions['C'].width = 25
ws_summary.column_dimensions['D'].width = 12
ws_summary.column_dimensions['E'].width = 25
ws_summary.column_dimensions['F'].width = 12
ws_summary.column_dimensions['G'].width = 25
ws_summary.column_dimensions['H'].width = 12
ws_summary.column_dimensions['I'].width = 25
ws_summary.column_dimensions['J'].width = 12
ws_summary.column_dimensions['K'].width = 30
ws_summary.column_dimensions['L'].width = 18


# ----------------- SHEET 2: SANDING RKA & RUP -----------------
ws_sanding = wb.create_sheet(title="Sanding RKA & RUP")
ws_sanding.views.sheetView[0].showGridLines = True

ws_sanding["A1"] = "PENYANDINGAN PAGU RKA DENGAN REALISASI RUP (TERUMUMKAN)"
ws_sanding["A1"].font = font_title
ws_sanding.merge_cells("A1:O1")

headers_sanding = [
    "No", "Program", "Kegiatan", "KRO (Output)", "RO (Sub-Output)", "Komponen", 
    "Key Otorisasi (MAK)", "Pagu RKA (A)", "Pagu Non-Pengadaan (B)", 
    "Target Pengadaan (C = A - B)", "Pagu RUP Terumumkan (D)", 
    "Selisih Pengadaan (C - D)", "Persentase (%)", "Status Evaluasi"
]

header_row = 4
for col_idx, h in enumerate(headers_sanding, start=1):
    cell = ws_sanding.cell(row=header_row, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_cell

def write_sanding_row(ws, row_idx, num_val, prog_code, keg_name, kro_name, ro_name, comp_name, comp_key,
                      pagu_rka, np_gaji_sum, target_pengadaan, rup_sum,
                      bg_fill, font_style):
    selisih_pengadaan = target_pengadaan - rup_sum
    
    if target_pengadaan == 0:
        if rup_sum == 0:
            pct = 1.0
            status_text = "Sesuai (Non-Pengadaan)"
            status_fill = fill_green
        else:
            pct = 9.99
            status_text = "Kelebihan Umumkan"
            status_fill = fill_red
    else:
        pct = rup_sum / target_pengadaan
        if pct == 1.0:
            status_text = "Sesuai (100%)"
            status_fill = fill_green
        elif pct > 0 and pct < 1.0:
            status_text = f"Parsial ({pct*100:.1f}%)"
            status_fill = fill_lvl1
        elif pct > 1.0:
            status_text = f"Kelebihan ({pct*100:.1f}%)"
            status_fill = fill_red
        else:
            status_text = "Belum Diumumkan"
            status_fill = fill_red

    ws.cell(row=row_idx, column=1, value=num_val).alignment = align_center
    ws.cell(row=row_idx, column=2, value=prog_code).alignment = align_center
    ws.cell(row=row_idx, column=3, value=keg_name).alignment = align_left
    ws.cell(row=row_idx, column=4, value=kro_name).alignment = align_left
    ws.cell(row=row_idx, column=5, value=ro_name).alignment = align_left
    ws.cell(row=row_idx, column=6, value=comp_name).alignment = align_left
    ws.cell(row=row_idx, column=7, value=comp_key).alignment = align_center
    
    ws.cell(row=row_idx, column=8, value=pagu_rka).number_format = '#,##0'
    ws.cell(row=row_idx, column=9, value=np_gaji_sum).number_format = '#,##0'
    ws.cell(row=row_idx, column=10, value=target_pengadaan).number_format = '#,##0'
    ws.cell(row=row_idx, column=11, value=rup_sum).number_format = '#,##0'
    ws.cell(row=row_idx, column=12, value=selisih_pengadaan).number_format = '#,##0'
    
    c_pct = ws.cell(row=row_idx, column=13, value=pct)
    c_pct.number_format = '0.0%'
    c_pct.alignment = align_right
    
    status_cell = ws.cell(row=row_idx, column=14, value=status_text)
    status_cell.alignment = align_center
    status_cell.fill = status_fill
    
    for col_c in range(1, 15):
        cell = ws.cell(row=row_idx, column=col_c)
        cell.border = border_cell
        cell.font = font_style
        if bg_fill:
            cell.fill = bg_fill

row_idx = 5
num = 1
total_rup_pagu = 0
total_target_pengadaan = 0
total_non_pengadaan = 0

font_sub = Font(name="Segoe UI", size=9, bold=True)
font_det = Font(name="Segoe UI", size=9)

for prog in rka_data:
    prog_code = prog["text"].split("]")[0].replace("[", "").strip()
    
    prog_rka = 0
    prog_np = 0
    prog_target = 0
    prog_rup = 0
    
    for keg in prog.get("kegiatans", []):
        keg_rka = 0
        keg_np = 0
        keg_target = 0
        keg_rup = 0
        
        for out in keg.get("outputs", []):
            out_rka = 0
            out_np = 0
            out_target = 0
            out_rup = 0
            
            suboutputs_list = out.get("suboutputs", [])
            if not suboutputs_list:
                suboutputs_list = [{ "id": 0, "name": out["name"], "code": "000", "komponens": out.get("komponens", []) }]
                
            for ro in suboutputs_list:
                ro_rka = 0
                ro_np = 0
                ro_target = 0
                ro_rup = 0
                
                for komp in ro.get("komponens", []):
                    comp_key = f"{prog_code}.{keg['code']}.{out['code']}.{ro['code']}.{komp['code']}"
                    
                    # Only sum RUP lines that are fully announced (A, FD, U checked)
                    rup_sum = sum(line["pagu"] for line in rup_all_lines if line["comp_key"] == comp_key and line["is_terumumkan"])
                    
                    # Hierarchical propagation of NP/Gaji flags
                    komp_np = komp_gj = False
                    subkomp_np = subkomp_gj = False
                    akun_np = akun_gj = False
                    
                    np_gaji_sum = 0
                    for r in komp.get("rows", []):
                        lvl = r["level"]
                        if lvl == 0:
                            komp_np = r["np_ch"]
                            komp_gj = r["gj_ch"]
                        elif lvl == 1:
                            subkomp_np = r["np_ch"]
                            subkomp_gj = r["gj_ch"]
                        elif lvl == 2:
                            akun_np = r["np_ch"]
                            akun_gj = r["gj_ch"]
                        elif lvl == 3:
                            is_np = r["np_ch"] or akun_np or subkomp_np or komp_np
                            is_gj = r["gj_ch"] or akun_gj or subkomp_gj or komp_gj
                            if is_np or is_gj:
                                np_gaji_sum += r["pagu"] if isinstance(r["pagu"], (int, float)) else 0
                    
                    target_pengadaan = komp["pagu"] - np_gaji_sum
                    
                    # Write detail row
                    write_sanding_row(
                        ws_sanding, row_idx, num, prog_code, keg["name"], out["name"], ro["name"], komp["name"], comp_key,
                        komp["pagu"], np_gaji_sum, target_pengadaan, rup_sum,
                        None, font_det
                    )
                    
                    ro_rka += komp["pagu"]
                    ro_np += np_gaji_sum
                    ro_target += target_pengadaan
                    ro_rup += rup_sum
                    
                    row_idx += 1
                    num += 1
                
                # Write Subtotal RO (Level 3 - Softest Gray)
                write_sanding_row(
                    ws_sanding, row_idx, "", prog_code, "", "", "", f"SUBTOTAL RO: {ro['name']}", f"{prog_code}.{keg['code']}.{out['code']}.{ro['code']}",
                    ro_rka, ro_np, ro_target, ro_rup,
                    fill_sub_ro, font_sub
                )
                
                out_rka += ro_rka
                out_np += ro_np
                out_target += ro_target
                out_rup += ro_rup
                
                row_idx += 1
            
            # Write Subtotal KRO (Level 2 - Light Gray F2F2F2)
            write_sanding_row(
                ws_sanding, row_idx, "", prog_code, "", "", "", f"SUBTOTAL KRO: {out['name']}", f"{prog_code}.{keg['code']}.{out['code']}",
                out_rka, out_np, out_target, out_rup,
                fill_lvl2, font_sub
            )
            
            keg_rka += out_rka
            keg_np += out_np
            keg_target += out_target
            keg_rup += out_rup
            
            row_idx += 1
            
        # Write Subtotal Kegiatan (Level 1 - Light Blue DCE6F1)
        write_sanding_row(
            ws_sanding, row_idx, "", prog_code, "", "", "", f"SUBTOTAL KEGIATAN: {keg['name']}", f"{prog_code}.{keg['code']}",
            keg_rka, keg_np, keg_target, keg_rup,
            fill_lvl1, font_sub
        )
        
        prog_rka += keg_rka
        prog_np += keg_np
        prog_target += keg_target
        prog_rup += keg_rup
        
        row_idx += 1
        
    # Write Subtotal Program (Level 0 - Light Blue B8CCE4)
    write_sanding_row(
        ws_sanding, row_idx, "", prog_code, "", "", "", f"SUBTOTAL PROGRAM: {prog['text'].split(']')[1].strip()}", f"{prog_code}",
        prog_rka, prog_np, prog_target, prog_rup,
        fill_lvl0, font_sub
    )
    
    total_non_pengadaan += prog_np
    total_target_pengadaan += prog_target
    total_rup_pagu += prog_rup
    
    row_idx += 1

# Total Row Sanding
ws_sanding.cell(row=row_idx, column=7, value="TOTAL").font = Font(name="Segoe UI", size=10, bold=True)
ws_sanding.cell(row=row_idx, column=7).alignment = align_right
ws_sanding.cell(row=row_idx, column=7).border = border_cell

for col_c in (8, 9, 10, 11, 12):
    if col_c == 8: val = total_satker_pagu
    elif col_c == 9: val = total_non_pengadaan
    elif col_c == 10: val = total_target_pengadaan
    elif col_c == 11: val = total_rup_pagu
    else: val = total_target_pengadaan - total_rup_pagu
    
    cell = ws_sanding.cell(row=row_idx, column=col_c, value=val)
    cell.font = Font(name="Segoe UI", size=10, bold=True)
    cell.number_format = '#,##0'
    cell.alignment = align_right
    cell.border = border_cell

tot_pct = (total_rup_pagu / total_target_pengadaan) if total_target_pengadaan > 0 else 0
c_tot_pct = ws_sanding.cell(row=row_idx, column=13, value=tot_pct)
c_tot_pct.font = Font(name="Segoe UI", size=10, bold=True)
c_tot_pct.number_format = '0.0%'
c_tot_pct.alignment = align_right
c_tot_pct.border = border_cell

# Card Summary Info
ws_sanding["A2"] = f"Total Pagu RKA: Rp {total_satker_pagu:,} | Target Pengadaan: Rp {total_target_pengadaan:,} | Terumumkan RUP: Rp {total_rup_pagu:,} ({tot_pct*100:.1f}%)"
ws_sanding["A2"].font = Font(name="Segoe UI", size=10, italic=True)
ws_sanding.merge_cells("A2:O2")

ws_sanding.column_dimensions['A'].width = 6
ws_sanding.column_dimensions['B'].width = 10
ws_sanding.column_dimensions['C'].width = 25
ws_sanding.column_dimensions['D'].width = 25
ws_sanding.column_dimensions['E'].width = 25
ws_sanding.column_dimensions['F'].width = 25
ws_sanding.column_dimensions['G'].width = 22
ws_sanding.column_dimensions['H'].width = 15
ws_sanding.column_dimensions['I'].width = 22
ws_sanding.column_dimensions['J'].width = 22
ws_sanding.column_dimensions['K'].width = 22
ws_sanding.column_dimensions['L'].width = 22
ws_sanding.column_dimensions['M'].width = 12
ws_sanding.column_dimensions['N'].width = 25


# ----------------- SHEET 3: DAFTAR PAKET RUP -----------------
ws_rup = wb.create_sheet(title="Daftar Paket RUP")
ws_rup.views.sheetView[0].showGridLines = True

ws_rup["A1"] = "DAFTAR PAKET PENYEDIA EXISTING YANG TERUMUMKAN (RUP)"
ws_rup["A1"].font = font_title
ws_rup.merge_cells("A1:L1")

headers_rup = [
    "No", "ID Paket", "Nama Kegiatan (RUP)", "Nama Paket", "Pagu RUP (Rp)",
    "Waktu Pemilihan", "Sumber Dana", "A (Draft PPK)", "FD (Final Draft PPK)",
    "U (Terumumkan KPA)", "Kode Otorisasi (MAK)", "Mata Anggaran Mapped"
]

header_row = 4
for col_idx, h in enumerate(headers_rup, start=1):
    cell = ws_rup.cell(row=header_row, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_cell

row_idx = 5
for idx, p in enumerate(rup_packets, start=1):
    ws_rup.cell(row=row_idx, column=1, value=idx).alignment = align_center
    ws_rup.cell(row=row_idx, column=2, value=p["id"]).alignment = align_center
    ws_rup.cell(row=row_idx, column=3, value=p["keg_name"]).alignment = align_left
    ws_rup.cell(row=row_idx, column=4, value=p["name"]).alignment = align_left
    
    c_pagu = ws_rup.cell(row=row_idx, column=5, value=p["pagu"])
    c_pagu.number_format = '#,##0'
    c_pagu.alignment = align_right
    
    ws_rup.cell(row=row_idx, column=6, value=p["waktu"]).alignment = align_center
    ws_rup.cell(row=row_idx, column=7, value=p["sumber_dana"]).alignment = align_center
    
    ws_rup.cell(row=row_idx, column=8, value="✓" if p["aktif"] else "").alignment = align_center
    ws_rup.cell(row=row_idx, column=9, value="✓" if p["fd"] else "").alignment = align_center
    ws_rup.cell(row=row_idx, column=10, value="✓" if p["umumkan"] else "").alignment = align_center
    
    ws_rup.cell(row=row_idx, column=11, value=p["mak"]).alignment = align_center
    
    packet_comp_keys = list(set(line["comp_key"] for line in rup_all_lines if line["packet_id"] == p["id"]))
    ws_rup.cell(row=row_idx, column=12, value=", ".join(packet_comp_keys)).alignment = align_center
    
    for col_c in range(1, 13):
        cell = ws_rup.cell(row=row_idx, column=col_c)
        cell.border = border_cell
        cell.font = Font(name="Segoe UI", size=9)
        
    row_idx += 1

# Total Row for RUP
ws_rup.cell(row=row_idx, column=4, value="TOTAL").font = Font(name="Segoe UI", size=10, bold=True)
ws_rup.cell(row=row_idx, column=4).alignment = align_right
ws_rup.cell(row=row_idx, column=4).border = border_cell

c_tot_rup = ws_rup.cell(row=row_idx, column=5, value=total_rup_pagu)
c_tot_rup.font = Font(name="Segoe UI", size=10, bold=True)
c_tot_rup.number_format = '#,##0'
c_tot_rup.alignment = align_right
c_tot_rup.border = border_cell

ws_rup.column_dimensions['A'].width = 5
ws_rup.column_dimensions['B'].width = 12
ws_rup.column_dimensions['C'].width = 25
ws_rup.column_dimensions['D'].width = 35
ws_rup.column_dimensions['E'].width = 15
ws_rup.column_dimensions['F'].width = 15
ws_rup.column_dimensions['G'].width = 12
ws_rup.column_dimensions['H'].width = 15
ws_rup.column_dimensions['I'].width = 18
ws_rup.column_dimensions['J'].width = 18
ws_rup.column_dimensions['K'].width = 30
ws_rup.column_dimensions['L'].width = 20


# ----------------- SHEET 3A: PAKET RUP TANPA SANDINGAN -----------------
ws_no_sanding = wb.create_sheet(title="Paket RUP Tanpa Sandingan")
ws_no_sanding.views.sheetView[0].showGridLines = True

ws_no_sanding["A1"] = "DAFTAR PAKET RUP YANG TIDAK MEMILIKI SANDINGAN DI RKA"
ws_no_sanding["A1"].font = font_title
ws_no_sanding.merge_cells("A1:F1")

headers_no_sanding = [
    "No", "ID Paket RUP", "Nama Paket RUP", "Kode MAK di RUP", "Pagu Paket (Rp)", "Keterangan / Alasan"
]

header_row = 4
for col_idx, h in enumerate(headers_no_sanding, start=1):
    cell = ws_no_sanding.cell(row=header_row, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_cell

# Gather all RKA detailed Akun keys (7-parts)
rka_detailed_keys = set()
for prog in rka_data:
    prog_code = prog["text"].split("]")[0].replace("[", "").strip()
    for keg in prog.get("kegiatans", []):
        for out in keg.get("outputs", []):
            suboutputs_list = out.get("suboutputs", [])
            if not suboutputs_list:
                suboutputs_list = [{ "id": 0, "name": out["name"], "code": "000", "komponens": out.get("komponens", []) }]
            for ro in suboutputs_list:
                for komp in ro.get("komponens", []):
                    comp_key = f"{prog_code}.{keg['code']}.{out['code']}.{ro['code']}.{komp['code']}"
                    
                    subkomp_code = ""
                    for r in komp.get("rows", []):
                        if r["level"] == 1:
                            subkomp_code = r["code"]
                        elif r["level"] == 2:
                            rka_detailed_keys.add(f"{comp_key}.{subkomp_code}.{r['code']}")

unmatched_rup_lines = []

# Check each line in RUP against 7-part RKA key
for line in rup_all_lines:
    if line["key"] not in rka_detailed_keys:
        unmatched_rup_lines.append({
            "packet_id": line["packet_id"],
            "packet_name": line["packet_name"],
            "mak": line["mak"],
            "pagu": line["pagu"],
            "reason": "Mata Anggaran (MAK) tidak ditemukan di RKA (Anggaran dihapus atau salah input)"
        })

# Check packets without detailed lines
for p in rup_packets:
    p_lines = [l for l in rup_all_lines if l["packet_id"] == p["id"]]
    if not p_lines:
        parsed = parse_mak(p["mak"])
        if parsed and parsed["key"] not in rka_detailed_keys:
            unmatched_rup_lines.append({
                "packet_id": p["id"],
                "packet_name": p["name"],
                "mak": p["mak"] if p["mak"] else "(Kosong)",
                "pagu": p["pagu"],
                "reason": "Mata Anggaran (MAK) tidak ditemukan di RKA (Anggaran dihapus atau salah input)"
            })
        elif not parsed:
            unmatched_rup_lines.append({
                "packet_id": p["id"],
                "packet_name": p["name"],
                "mak": p["mak"] if p["mak"] else "(Kosong)",
                "pagu": p["pagu"],
                "reason": "Format Kode MAK di RUP tidak valid atau kosong"
            })

# Remove duplicates
seen = set()
unique_unmatched = []
for item in unmatched_rup_lines:
    key = (item["packet_id"], item["mak"])
    if key not in seen:
        seen.add(key)
        unique_unmatched.append(item)

row_idx = 5
for idx, item in enumerate(unique_unmatched, start=1):
    ws_no_sanding.cell(row=row_idx, column=1, value=idx).alignment = align_center
    ws_no_sanding.cell(row=row_idx, column=2, value=item["packet_id"]).alignment = align_center
    ws_no_sanding.cell(row=row_idx, column=3, value=item["packet_name"]).alignment = align_left
    ws_no_sanding.cell(row=row_idx, column=4, value=item["mak"]).alignment = align_center
    
    c_pagu = ws_no_sanding.cell(row=row_idx, column=5, value=item["pagu"])
    c_pagu.number_format = '#,##0'
    c_pagu.alignment = align_right
    
    ws_no_sanding.cell(row=row_idx, column=6, value=item["reason"]).alignment = align_left
    
    for col_c in range(1, 7):
        cell = ws_no_sanding.cell(row=row_idx, column=col_c)
        cell.border = border_cell
        cell.font = Font(name="Segoe UI", size=9)
        cell.fill = fill_red
        
    row_idx += 1

# Set widths
ws_no_sanding.column_dimensions['A'].width = 5
ws_no_sanding.column_dimensions['B'].width = 15
ws_no_sanding.column_dimensions['C'].width = 35
ws_no_sanding.column_dimensions['D'].width = 30
ws_no_sanding.column_dimensions['E'].width = 18
ws_no_sanding.column_dimensions['F'].width = 45


# ----------------- SHEET 4 & 5: DETAIL PER PROGRAM -----------------
for prog in rka_data:
    prog_code = prog["text"].split("]")[0].replace("[", "").strip()
    sheet_name = f"Detail - {prog_code}"
    ws = wb.create_sheet(title=sheet_name)
    ws.views.sheetView[0].showGridLines = True
    
    # Title
    ws["A1"] = f"DETAIL RENCANA KERJA ANGGARAN (RKA) - PROGRAM {prog_code}"
    ws["A1"].font = font_title
    ws.merge_cells("A1:S1")
    
    ws["A3"] = "Program:"
    ws["A3"].font = font_meta_label
    ws["B3"] = prog["text"]
    ws["B3"].font = font_meta_val
    ws.merge_cells("B3:S3")
    
    headers_det = [
        "Kegiatan", "KRO (Output)", "RO (Sub-Output)", "Komponen", "Kode (P/K/O/SO/K/SK/A/D)",
        "Uraian", "Uraian Sebelum Revisi", "Pagu RKA", "Pagu Sebelum Revisi",
        "P", "S", "Multiyears", "NP", "Gaji",
        "ID Paket RUP", "Nama Paket RUP", "Pagu RUP (Announced)", "Selisih Pengadaan", "Rencana Pemilihan"
    ]
    
    # Write Headers
    header_row = 5
    for col_idx, h in enumerate(headers_det, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_cell
        
    curr_row = 6
    matched_rup_indices = set()
    
    for keg in prog.get("kegiatans", []):
        keg_text = f"[{keg['code']}] {keg['name']}"
        for out in keg.get("outputs", []):
            kro_text = f"[{out['code']}] {out['name']}"
            
            suboutputs_list = out.get("suboutputs", [])
            if not suboutputs_list:
                suboutputs_list = [{ "id": 0, "name": out["name"], "code": "000", "komponens": out.get("komponens", []) }]
                
            for ro in suboutputs_list:
                ro_text = f"[{ro['code']}] {ro['name']}"
                
                for komp in ro.get("komponens", []):
                    komp_text = f"[{komp['code']}] {komp['name']}"
                    
                    comp_key = f"{prog_code}.{keg['code']}.{out['code']}.{ro['code']}.{komp['code']}"
                    comp_rup_lines = [line for line in rup_all_lines if line["comp_key"] == comp_key]
                    
                    start_row = curr_row
                    
                    # --- PRE-CALCULATE target_pagu FOR THIS COMPONENT WITH PROPAGATION ---
                    rows_target = {}
                    rows_is_np = {}
                    rows_is_gj = {}
                    
                    komp_np = komp_gj = False
                    subkomp_np = subkomp_gj = False
                    akun_np = akun_gj = False
                    
                    for r_idx, r in enumerate(komp.get("rows", [])):
                        lvl = r["level"]
                        pagu = r["pagu"] if isinstance(r["pagu"], (int, float)) else 0
                        
                        if lvl == 0:
                            komp_np = r["np_ch"]
                            komp_gj = r["gj_ch"]
                        elif lvl == 1:
                            subkomp_np = r["np_ch"]
                            subkomp_gj = r["gj_ch"]
                        elif lvl == 2:
                            akun_np = r["np_ch"]
                            akun_gj = r["gj_ch"]
                        elif lvl == 3:
                            is_np = r["np_ch"] or akun_np or subkomp_np or komp_np
                            is_gj = r["gj_ch"] or akun_gj or subkomp_gj or komp_gj
                            rows_is_np[r_idx] = is_np
                            rows_is_gj[r_idx] = is_gj
                            rows_target[r_idx] = 0 if (is_np or is_gj) else pagu
                    
                    # Calculate level 2 (Akun) target pagus
                    curr_akun_idx = None
                    for r_idx, r in enumerate(komp.get("rows", [])):
                        lvl = r["level"]
                        if lvl == 2:
                            curr_akun_idx = r_idx
                            rows_target[r_idx] = 0
                        elif lvl == 3 and curr_akun_idx is not None:
                            rows_target[curr_akun_idx] += rows_target[r_idx]
                            
                    # Calculate level 1 (Sub-komponen) target pagus
                    curr_sub_idx = None
                    for r_idx, r in enumerate(komp.get("rows", [])):
                        lvl = r["level"]
                        if lvl == 1:
                            curr_sub_idx = r_idx
                            rows_target[r_idx] = 0
                        elif lvl == 3 and curr_sub_idx is not None:
                            rows_target[curr_sub_idx] += rows_target[r_idx]
                            
                    # Calculate level 0 (Komponen) target pagus
                    curr_komp_idx = None
                    for r_idx, r in enumerate(komp.get("rows", [])):
                        lvl = r["level"]
                        if lvl == 0:
                            curr_komp_idx = r_idx
                            rows_target[r_idx] = 0
                        elif lvl == 3 and curr_komp_idx is not None:
                            rows_target[curr_komp_idx] += rows_target[r_idx]
                    
                    # --- DETAILED MATCHING FOR LEVEL 3 DETAIL ROWS (PROPORTIONAL ALLOCATION) ---
                    detail_row_objects = []
                    curr_subkomp = ""
                    curr_akun = ""
                    for r_idx, r in enumerate(komp.get("rows", [])):
                        lvl = r["level"]
                        if lvl == 1:
                            curr_subkomp = r["code"]
                        elif lvl == 2:
                            curr_akun = r["code"]
                        elif lvl == 3:
                            is_np = rows_is_np.get(r_idx, False)
                            is_gj = rows_is_gj.get(r_idx, False)
                            if not (is_np or is_gj):
                                detail_row_objects.append({
                                    "r_idx": r_idx,
                                    "pagu": r["pagu"] if isinstance(r["pagu"], (int, float)) else 0,
                                    "key": f"{comp_key}.{curr_subkomp}.{curr_akun}",
                                    "allocated_pagu": 0,
                                    "matched_ids": [],
                                    "matched_names": [],
                                    "is_draft_match": False
                                })
                                
                    # Group RUP lines by their Akun key and separate into Terumumkan and Draft pools
                    rup_pools_terumumkan = []
                    rup_pools_draft = []
                    for line in comp_rup_lines:
                        item = {
                            "packet_id": line["packet_id"],
                            "packet_name": line["packet_name"],
                            "pagu": line["pagu"],
                            "remaining_pagu": line["pagu"],
                            "key": line["key"]
                        }
                        if line["is_terumumkan"]:
                            rup_pools_terumumkan.append(item)
                        else:
                            rup_pools_draft.append(item)
                            
                    # Perform Proportional Allocation for each unique Akun key
                    unique_keys = set(d["key"] for d in detail_row_objects)
                    for k in unique_keys:
                        k_rka_rows = [d for d in detail_row_objects if d["key"] == k]
                        k_rup_pools_ter = [p for p in rup_pools_terumumkan if p["key"] == k]
                        k_rup_pools_drf = [p for p in rup_pools_draft if p["key"] == k]
                        
                        # 1. First Pass: Allocate Terumumkan RUP
                        for d_obj in k_rka_rows:
                            needed = d_obj["pagu"]
                            for pool in k_rup_pools_ter:
                                if pool["remaining_pagu"] > 0 and needed > 0:
                                    amount = min(needed, pool["remaining_pagu"])
                                    pool["remaining_pagu"] -= amount
                                    d_obj["allocated_pagu"] += amount
                                    needed -= amount
                                    
                                    if pool["packet_id"] not in d_obj["matched_ids"]:
                                        d_obj["matched_ids"].append(pool["packet_id"])
                                    if pool["packet_name"] not in d_obj["matched_names"]:
                                        d_obj["matched_names"].append(pool["packet_name"])
                                        
                        # Leftover Terumumkan RUP goes to the first RKA row
                        leftover = sum(p["remaining_pagu"] for p in k_rup_pools_ter)
                        if leftover > 0 and k_rka_rows:
                            first_row = k_rka_rows[0]
                            first_row["allocated_pagu"] += leftover
                            for pool in k_rup_pools_ter:
                                if pool["remaining_pagu"] > 0:
                                    if pool["packet_id"] not in first_row["matched_ids"]:
                                        first_row["matched_ids"].append(pool["packet_id"])
                                    if pool["packet_name"] not in first_row["matched_names"]:
                                        first_row["matched_names"].append(pool["packet_name"])
                                    pool["remaining_pagu"] = 0
                                    
                        # 2. Second Pass: For rows that remain unmatched/partially matched, link to Draft RUPs
                        for d_obj in k_rka_rows:
                            if d_obj["allocated_pagu"] < d_obj["pagu"]:
                                for pool in k_rup_pools_drf:
                                    if pool["remaining_pagu"] > 0:
                                        d_obj["is_draft_match"] = True
                                        drf_id = f"[DRAFT/BATAL] {pool['packet_id']}"
                                        drf_name = f"[Draft/Batal] {pool['packet_name']}"
                                        
                                        if drf_id not in d_obj["matched_ids"]:
                                            d_obj["matched_ids"].append(drf_id)
                                        if drf_name not in d_obj["matched_names"]:
                                            d_obj["matched_names"].append(drf_name)
                                        pool["remaining_pagu"] = 0

                    # --- ITERATE AND WRITE ROWS ---
                    subkomp_code = ""
                    akun_code = ""
                    
                    for row_idx_in_comp, row_data in enumerate(komp.get("rows", [])):
                        level = row_data["level"]
                        code = row_data["code"]
                        desc = row_data["desc"]
                        desc_prev = row_data["descPrev"]
                        pagu = row_data["pagu"]
                        pagu_prev = row_data["paguPrev"]
                        
                        p_ch = row_data["p_ch"]
                        s_ch = row_data["s_ch"]
                        my_ch = row_data["my_ch"]
                        np_ch = row_data["np_ch"]
                        gj_ch = row_data["gj_ch"]
                        
                        if level == 1:
                            subkomp_code = code
                        elif level == 2:
                            akun_code = code
                        
                        ws.cell(row=curr_row, column=5, value=code)
                        ws.cell(row=curr_row, column=6, value=desc)
                        ws.cell(row=curr_row, column=7, value=desc_prev)
                        
                        if isinstance(pagu, (int, float)):
                            ws.cell(row=curr_row, column=8, value=pagu).number_format = '#,##0'
                        else:
                            ws.cell(row=curr_row, column=8, value=pagu)
                            
                        if isinstance(pagu_prev, (int, float)):
                            ws.cell(row=curr_row, column=9, value=pagu_prev).number_format = '#,##0'
                        else:
                            ws.cell(row=curr_row, column=9, value=pagu_prev)
                            
                        ws.cell(row=curr_row, column=10, value="✓" if p_ch else "")
                        ws.cell(row=curr_row, column=11, value="✓" if s_ch else "")
                        ws.cell(row=curr_row, column=12, value="✓" if my_ch else "")
                        ws.cell(row=curr_row, column=13, value="✓" if np_ch else "")
                        ws.cell(row=curr_row, column=14, value="✓" if gj_ch else "")
                        
                        matched_pkt_id = ""
                        matched_pkt_name = ""
                        rup_pagu_val = 0
                        is_np_gaji = False
                        is_draft = False
                        matched_pkt_waktu = ""
                        
                        if level == 0:
                            rup_pagu_val = sum(line["pagu"] for line in comp_rup_lines if line["is_terumumkan"])
                        elif level == 1:
                            rup_pagu_val = sum(line["pagu"] for line in comp_rup_lines if line["subkomp"] == subkomp_code and line["is_terumumkan"])
                        elif level == 2:
                            akun_key = f"{prog_code}.{keg['code']}.{out['code']}.{ro['code']}.{komp['code']}.{subkomp_code}.{code}"
                            rup_pagu_val = sum(line["pagu"] for line in comp_rup_lines if line["key"] == akun_key and line["is_terumumkan"])
                        elif level == 3:
                            is_np = rows_is_np.get(row_idx_in_comp, False)
                            is_gj = rows_is_gj.get(row_idx_in_comp, False)
                            is_np_gaji = is_np or is_gj
                            
                            if is_np:
                                matched_pkt_id = "NP"
                                matched_pkt_name = "Non-Pengadaan (Honor/Perdin/Uang Makan/PJLP/PPPK Paruh Waktu)"
                            elif is_gj:
                                matched_pkt_id = "Gaji"
                                matched_pkt_name = "Gaji & Tunjangan Pegawai"
                            else:
                                d_obj = next((d for d in detail_row_objects if d["r_idx"] == row_idx_in_comp), None)
                                if d_obj and d_obj["matched_ids"]:
                                    matched_pkt_id = ", ".join(str(i) for i in d_obj["matched_ids"])
                                    matched_pkt_name = ", ".join(d_obj["matched_names"])
                                    rup_pagu_val = d_obj["allocated_pagu"]
                                    is_draft = d_obj["is_draft_match"]
                                    
                                    matched_times = []
                                    for p_id in d_obj["matched_ids"]:
                                        clean_id = str(p_id).replace("[DRAFT/BATAL] ", "").strip()
                                        pkt = next((p for p in rup_packets if str(p["id"]) == clean_id), None)
                                        if pkt and pkt.get("waktu") and pkt["waktu"] not in matched_times:
                                            matched_times.append(pkt["waktu"])
                                    matched_pkt_waktu = ", ".join(matched_times)
                        
                        ws.cell(row=curr_row, column=15, value=matched_pkt_id)
                        ws.cell(row=curr_row, column=16, value=matched_pkt_name)
                        ws.cell(row=curr_row, column=19, value=matched_pkt_waktu)
                        
                        target_pagu = rows_target.get(row_idx_in_comp, 0)
                        
                        if is_np_gaji:
                            c_rp = ws.cell(row=curr_row, column=17, value=0)
                            c_rp.number_format = '#,##0'
                            c_df = ws.cell(row=curr_row, column=18, value=0)
                            c_df.number_format = '#,##0'
                        else:
                            c_rp = ws.cell(row=curr_row, column=17, value=rup_pagu_val)
                            c_rp.number_format = '#,##0'
                            c_df = ws.cell(row=curr_row, column=18, value=target_pagu - rup_pagu_val)
                            c_df.number_format = '#,##0'
                        
                        if level == 0:
                            row_fill = fill_lvl0
                            row_font = font_lvl0
                        elif level == 1:
                            row_fill = fill_lvl1
                            row_font = font_lvl1
                        elif level == 2:
                            row_fill = fill_lvl2
                            row_font = font_lvl2
                        else:
                            row_fill = fill_lvl3
                            row_font = font_lvl3
                        
                        # 1. Apply default cell layout and fill styles
                        for col_idx in range(5, 20):
                            cell = ws.cell(row=curr_row, column=col_idx)
                            cell.fill = row_fill
                            if col_idx < 15:
                                cell.font = row_font
                            cell.border = border_cell
                            
                            if col_idx == 5:
                                cell.alignment = align_center
                            elif col_idx in (6, 7):
                                cell.alignment = align_with_indent(level)
                            elif col_idx in (8, 9):
                                cell.alignment = align_right
                            elif col_idx in (10, 11, 12, 13, 14, 19):
                                cell.alignment = align_center
                            elif col_idx in (15, 17, 18):
                                cell.alignment = align_center
                            elif col_idx == 16:
                                cell.alignment = align_left

                        # 2. Apply level 3 specific font/fill colors (warning/match/np overrides)
                        if level == 3:
                            if is_draft:
                                for col_idx in (15, 16, 17, 19):
                                    c = ws.cell(row=curr_row, column=col_idx)
                                    c.fill = fill_draft
                                    c.font = font_draft
                                ws.cell(row=curr_row, column=6).font = font_lvl3
                            elif matched_pkt_id and matched_pkt_id not in ("NP", "Gaji"):
                                ws.cell(row=curr_row, column=6).font = font_match
                                for col_idx in (15, 16, 17, 19):
                                    ws.cell(row=curr_row, column=col_idx).font = font_match
                            elif is_np_gaji:
                                ws.cell(row=curr_row, column=6).font = font_np
                                for col_idx in range(15, 20):
                                    ws.cell(row=curr_row, column=col_idx).font = font_np
                        
                        curr_row += 1
                    
                    end_row = curr_row - 1
                    if end_row >= start_row:
                        ws.cell(row=start_row, column=1, value=keg_text)
                        ws.cell(row=start_row, column=2, value=kro_text)
                        ws.cell(row=start_row, column=3, value=ro_text)
                        ws.cell(row=start_row, column=4, value=komp_text)
                        
                        for r in range(start_row, end_row + 1):
                            for c in (1, 2, 3, 4):
                                cell = ws.cell(row=r, column=c)
                                cell.font = font_hierarchy
                                cell.alignment = align_hierarchy
                                cell.border = border_cell
                                cell.fill = PatternFill(start_color="FAFBFD", end_color="FAFBFD", fill_type="solid")
                        
                        if end_row > start_row:
                            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                            ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
                            ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)
                            ws.merge_cells(start_row=start_row, start_column=4, end_row=end_row, end_column=4)
                    
                    curr_row += 2

    # Freeze rows 1 to 5
    ws.freeze_panes = "A6"
    
    # Auto Filter
    ws.auto_filter.ref = f"E5:S{curr_row - 1}"
    
    # Dimensions
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 45
    ws.column_dimensions['G'].width = 45
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    for col_c in range(10, 15):
        col_letter = get_column_letter(col_c)
        ws.column_dimensions[col_letter].width = 8
    ws.column_dimensions['O'].width = 15
    ws.column_dimensions['P'].width = 30
    ws.column_dimensions['Q'].width = 20
    ws.column_dimensions['R'].width = 20
    ws.column_dimensions['S'].width = 18

# Save
try:
    wb.save("rekap_rka_lengkap.xlsx")
    print("Consolidated Excel with unique RO keys and hierarchical subtotals updated successfully!")
except PermissionError:
    import time
    alt_name = f"rekap_rka_lengkap_{int(time.time())}.xlsx"
    wb.save(alt_name)
    print(f"Permission denied on original file. Saved as alternative: {alt_name}")
