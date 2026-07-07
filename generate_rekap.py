import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Rekap RKA"
ws.views.sheetView[0].showGridLines = True

# Colors
DARK_BLUE = "1F497D"
WHITE = "FFFFFF"
LIGHT_BLUE_LVL0 = "B8CCE4"  # B8CCE4
LIGHT_BLUE_LVL1 = "DCE6F1"  # DCE6F1
LIGHT_GRAY_LVL2 = "F2F2F2"  # F2F2F2
WHITE_COLOR = "FFFFFF"

# Fonts
font_title = Font(name="Segoe UI", size=16, bold=True, color="1F497D")
font_meta_label = Font(name="Segoe UI", size=10, bold=True)
font_meta_val = Font(name="Segoe UI", size=10)
font_header = Font(name="Segoe UI", size=10, bold=True, color=WHITE)
font_lvl0 = Font(name="Segoe UI", size=10, bold=True)
font_lvl1 = Font(name="Segoe UI", size=10, bold=True)
font_lvl2 = Font(name="Segoe UI", size=10, italic=False)
font_lvl3 = Font(name="Segoe UI", size=9, italic=True, color="595959")

# Fills
fill_header = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
fill_lvl0 = PatternFill(start_color=LIGHT_BLUE_LVL0, end_color=LIGHT_BLUE_LVL0, fill_type="solid")
fill_lvl1 = PatternFill(start_color=LIGHT_BLUE_LVL1, end_color=LIGHT_BLUE_LVL1, fill_type="solid")
fill_lvl2 = PatternFill(start_color=LIGHT_GRAY_LVL2, end_color=LIGHT_GRAY_LVL2, fill_type="solid")
fill_lvl3 = PatternFill(start_color=WHITE_COLOR, end_color=WHITE_COLOR, fill_type="solid")

# Alignments
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

def align_with_indent(level):
    return Alignment(horizontal="left", vertical="center", indent=level * 2)

# Borders
thin_side = Side(border_style="thin", color="D3D3D3")
border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# Title Block
ws["A1"] = "REKAPITULASI RENCANA KERJA ANGGARAN (RKA) TAHUN 2026"
ws["A1"].font = font_title
ws.merge_cells("A1:J1")

meta = [
    ("Program", "[DL] Program Pendidikan dan Pelatihan Vokasi"),
    ("Kegiatan", "[2378] Dukungan Manajemen Internal Lingkup Badan Penyuluhan dan Pengembangan Sumber Daya Manusia Kelautan dan Perikanan"),
    ("KRO", "[FAN] Pemenuhan Prioritas Direktif Presiden"),
    ("Komponen", "[101] Penyusunan Peraturan Kelautan dan Perikanan"),
    ("Pengguna", "ppkmujib-218")
]

for idx, (label, val) in enumerate(meta, start=3):
    ws.cell(row=idx, column=1, value=label).font = font_meta_label
    ws.cell(row=idx, column=2, value=val).font = font_meta_val
    ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=10)

# Table Headers
headers = [
    "Kode (P/K/O/SO/K/SK/A/D)",
    "Uraian",
    "Uraian Sebelum Revisi",
    "Pagu",
    "Pagu Sebelum Revisi",
    "P",
    "S",
    "Multiyears",
    "NP",
    "Gaji"
]

header_row = 9
for col_idx, h in enumerate(headers, start=1):
    cell = ws.cell(row=header_row, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_cell

# Data Rows
data = [
    (0, "101", "Penyusunan Peraturan Kelautan dan Perikanan", "Penyusunan Peraturan Kelautan dan Perikanan", 71352000, 0, False, False, False, False, False),
    (1, "JA", "Penyusunan Peraturan Menteri KP", "Penyusunan Peraturan Menteri KP", 36193000, 0, False, False, False, False, False),
    (2, "521211", "Belanja Bahan", "Belanja Bahan", 9473000, 0, True, False, False, False, False),
    (3, "1", "ATK", "ATK", 0, 0, False, False, False, False, False),
    (3, "01", "Bahan Komputer", "Bahan Komputer", 917000, 0, False, False, False, False, False),
    (3, "2", "Bahan Komputer", "Bahan Komputer", 0, 0, False, False, False, False, False),
    (3, "02", "Makan Rapat Biasa", "Makan Rapat Biasa", 5700000, 0, False, False, False, False, False),
    (3, "3", "Makan Rapat Biasa", "Makan Rapat Biasa", 0, 0, False, False, False, False, False),
    (3, "03", "Snack Rapat Biasa", "Snack Rapat Biasa", 2400000, 0, False, False, False, False, False),
    (3, "4", "Snack Rapat Biasa", "Snack Rapat Biasa", 0, 0, False, False, False, False, False),
    (3, "04", "Fotokopi dan penggandaan", "Fotokopi dan penggandaan", 456000, 0, False, False, False, False, False),
    (3, "5", "Fotokopi dan penggandaan", "Fotokopi dan penggandaan", 0, 0, False, False, False, False, False),
    (3, "6", "Pencetakan Banner", "Pencetakan Banner", 0, 0, False, False, False, False, False),
    (2, "522151", "Belanja Jasa Profesi", "Belanja Jasa Profesi", 0, 0, False, False, False, True, False),
    (3, "1", "Honor Narasumber pejabat/setara eselon II kebawah", "Honor Narasumber pejabat/setara eselon II kebawah", 0, 0, False, False, False, False, False),
    (2, "524111", "Belanja Perjalanan Dinas Biasa", "Belanja Perjalanan Dinas Biasa", 20000000, 0, False, False, False, True, False),
    (3, "01", "Uang Harian", "-", 6020000, "-", False, False, False, False, False),
    (3, "01", "Penginapan", "Penginapan", 0, 0, False, False, False, False, False),
    (3, "02", "Transport", "-", 13980000, "-", False, False, False, False, False),
    (3, "02", "Uang Harian", "Uang Harian", 0, 0, False, False, False, False, False),
    (3, "03", "Transport", "Transport", 0, 0, False, False, False, False, False),
    (2, "524113", "Belanja Perjalanan Dinas Dalam Kota", "Belanja Perjalanan Dinas Dalam Kota", 1360000, 0, False, False, False, True, False),
    (3, "01", "Transport Lokal", "Transport Lokal", 1360000, 0, False, False, False, False, False),
    (2, "524119", "Belanja Perjalanan Dinas Paket Meeting Luar Kota", "Belanja Perjalanan Dinas Paket Meeting Luar Kota", 5360000, 0, False, False, False, True, False),
    (3, "01", "Perjalanan Peserta", "Perjalanan Peserta", 4320000, 0, False, False, False, False, False),
    (3, "02", "Uang Harian Peserta", "Uang Harian Peserta", 1040000, 0, False, False, False, False, False),
    (3, "3", "Paket Meeting Peserta", "Paket Meeting Peserta", 0, 0, False, False, False, False, False),
    (1, "JB", "Penyusunan Keputusan Menteri KP", "Penyusunan Keputusan Menteri KP", 35159000, 0, False, False, False, False, False),
    (2, "521211", "Belanja Bahan", "Belanja Bahan", 8099000, 0, True, False, False, False, False),
    (3, "1", "ATK", "ATK", 0, 0, False, False, False, False, False),
    (3, "01", "Bahan Komputer", "Bahan Komputer", 567000, 0, False, False, False, False, False),
    (3, "2", "Bahan Komputer", "Bahan Komputer", 0, 0, False, False, False, False, False),
    (3, "02", "ATK", "ATK", 319000, 0, False, False, False, False, False),
    (3, "03", "Makan Rapat Biasa", "Makan Rapat Biasa", 4846000, 0, False, False, False, False, False),
    (3, "04", "Snack Rapat Biasa", "Snack Rapat Biasa", 2040000, 0, False, False, False, False, False),
    (3, "05", "Fotokopi dan penggandaan", "Fotokopi dan penggandaan", 327000, 0, False, False, False, False, False),
    (3, "6", "Pencetakan Banner", "Pencetakan Banner", 0, 0, False, False, False, False, False),
    (2, "522151", "Belanja Jasa Profesi", "Belanja Jasa Profesi", 0, 0, False, False, False, True, False),
    (3, "1", "Honor Narasumber pejabat/setara eselon II kebawah", "Honor Narasumber pejabat/setara eselon II kebawah", 0, 0, False, False, False, False, False),
    (2, "524111", "Belanja Perjalanan Dinas Biasa", "Belanja Perjalanan Dinas Biasa", 12000000, 0, False, False, False, True, False),
    (3, "01", "Penginapan", "Penginapan", 0, 0, False, False, False, False, False),
    (3, "01", "Transport", "-", 12000000, "-", False, False, False, False, False),
    (3, "02", "Uang Harian", "Uang Harian", 0, 0, False, False, False, False, False),
    (3, "03", "Transport", "Transport", 0, 0, False, False, False, False, False),
    (2, "524113", "Belanja Perjalanan Dinas Dalam Kota", "Belanja Perjalanan Dinas Dalam Kota", 1700000, 0, False, False, False, True, False),
    (3, "01", "Transport Lokal", "Transport Lokal", 1700000, 0, False, False, False, False, False),
    (2, "524119", "Belanja Perjalanan Dinas Paket Meeting Luar Kota", "Belanja Perjalanan Dinas Paket Meeting Luar Kota", 13360000, 0, False, False, False, True, False),
    (3, "01", "Perjalanan Peserta", "Perjalanan Peserta", 4320000, 0, False, False, False, False, False),
    (3, "02", "Uang Harian Peserta", "Uang Harian Peserta", 1040000, 0, False, False, False, False, False),
    (3, "03", "Paket Meeting Peserta", "Paket Meeting Peserta", 8000000, 0, False, False, False, False, False)
]

current_row = 10
for item in data:
    level, code, desc, desc_prev, pagu, pagu_prev, p_ch, s_ch, my_ch, np_ch, gj_ch = item
    
    # Write values
    c_code = ws.cell(row=current_row, column=1, value=code)
    c_desc = ws.cell(row=current_row, column=2, value=desc)
    c_desc_prev = ws.cell(row=current_row, column=3, value=desc_prev)
    
    # Handle numeric values for pagu and pagu_prev
    if isinstance(pagu, (int, float)):
        c_pagu = ws.cell(row=current_row, column=4, value=pagu)
        c_pagu.number_format = '#,##0'
    else:
        c_pagu = ws.cell(row=current_row, column=4, value=pagu)
        
    if isinstance(pagu_prev, (int, float)):
        c_pagu_prev = ws.cell(row=current_row, column=5, value=pagu_prev)
        c_pagu_prev.number_format = '#,##0'
    else:
        c_pagu_prev = ws.cell(row=current_row, column=5, value=pagu_prev)

    # Checkboxes
    ws.cell(row=current_row, column=6, value="✓" if p_ch else "")
    ws.cell(row=current_row, column=7, value="✓" if s_ch else "")
    ws.cell(row=current_row, column=8, value="✓" if my_ch else "")
    ws.cell(row=current_row, column=9, value="✓" if np_ch else "")
    ws.cell(row=current_row, column=10, value="✓" if gj_ch else "")

    # Formatting based on level
    if level == 0:
        row_fill = fill_lvl0
        row_font = font_lvl0
    elif level == 1:
        row_fill = fill_lvl1
        row_font = font_lvl1
    elif level == 2:
        row_fill = fill_lvl2
        row_font = font_lvl2
    else:
        row_fill = fill_lvl3
        row_font = font_lvl3
        
    for col_idx in range(1, 11):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.fill = row_fill
        cell.font = row_font
        cell.border = border_cell
        
        # Alignments
        if col_idx == 1:
            cell.alignment = align_center
        elif col_idx == 2:
            cell.alignment = align_with_indent(level)
        elif col_idx == 3:
            cell.alignment = align_with_indent(level)
        elif col_idx in (4, 5):
            cell.alignment = align_right
        else:
            cell.alignment = align_center

    current_row += 1

# Auto-adjust column widths
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        # Ignore merged title row A1 and meta block
        if cell.row < 9:
            continue
        val_str = str(cell.value or '')
        if len(val_str) > max_len:
            max_len = len(val_str)
    # Add a bit of padding and set width
    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

# Specific manual adjustments
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 50
ws.column_dimensions['A'].width = 15

wb.save("rekap_rka_101.xlsx")
print("Excel generated successfully!")
