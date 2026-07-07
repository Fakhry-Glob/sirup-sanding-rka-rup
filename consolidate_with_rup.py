import json
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Paths
rka_txt_path = r"C:\Users\user\.gemini\antigravity\brain\b120cd73-e77b-45ef-a520-5f577420490e\.system_generated\steps\104\output.txt"
rup_txt_path = r"C:\Users\user\.gemini\antigravity\brain\b120cd73-e77b-45ef-a520-5f577420490e\.system_generated\steps\160\output.txt"

# 1. Load RKA Data
with open(rka_txt_path, "r", encoding="utf-8") as f:
    rka_content = f.read()

rka_match = re.search(r"(\[.*\]|\{.*\})", rka_content.split("### Ran Playwright code")[0], re.DOTALL)
if not rka_match:
    raise ValueError("RKA JSON not found")
rka_data = json.loads(rka_match.group(1))

# 2. Load RUP Data
with open(rup_txt_path, "r", encoding="utf-8") as f:
    rup_content = f.read()

rup_match = re.search(r"(\[.*\]|\{.*\})", rup_content.split("### Ran Playwright code")[0], re.DOTALL)
if not rup_match:
    rup_match = re.search(r"(\[.*\]|\{.*\})", rup_content, re.DOTALL)

if not rup_match:
    raise ValueError("RUP JSON not found")

rup_json = json.loads(rup_match.group(1))
rup_rows = rup_json.get("rows", [])

# 3. Process RUP Packets and group by Component Key
rup_packets = []
rup_by_comp = {}

for row in rup_rows:
    packet_id = row[0]
    keg_name = row[1]
    packet_name = row[2]
    pagu = int(row[3]) if str(row[3]).isdigit() else 0
    waktu = row[4]
    sumber_dana = row[5]
    aktif = row[6] == "true"  # A (Draft PPK)
    fd = row[7] == "true"     # FD (Final Draft PPK)
    umumkan = row[8] == "true" # U (Terumumkan KPA)
    mak_code = row[17] if len(row) > 17 else ""
    
    mak_parts = mak_code.split(".")
    prog_code = ""
    keg_code = ""
    out_code = ""
    komp_code = ""
    comp_key = ""
    
    if len(mak_parts) >= 7:
        prog_code = mak_parts[2]
        keg_code = mak_parts[3]
        out_code = mak_parts[4]
        komp_code = mak_parts[6]
        comp_key = f"{prog_code}.{keg_code}.{out_code}.{komp_code}"
        
    packet_info = {
        "id": packet_id,
        "keg_name": keg_name,
        "name": packet_name,
        "pagu": pagu,
        "waktu": waktu,
        "sumber_dana": sumber_dana,
        "aktif": aktif,
        "fd": fd,
        "umumkan": umumkan,
        "mak": mak_code,
        "comp_key": comp_key,
        "prog_code": prog_code,
        "keg_code": keg_code,
        "out_code": out_code,
        "komp_code": komp_code
    }
    
    rup_packets.append(packet_info)
    
    if comp_key:
        if comp_key not in rup_by_comp:
            rup_by_comp[comp_key] = []
        rup_by_comp[comp_key].append(packet_info)

# Create workbook
wb = openpyxl.Workbook()

# Styles
DARK_BLUE = "1F497D"
WHITE = "FFFFFF"
LIGHT_BLUE_LVL0 = "B8CCE4"
LIGHT_BLUE_LVL1 = "DCE6F1"
LIGHT_GRAY_LVL2 = "F2F2F2"
LIGHT_GREEN = "E2EFDA"
LIGHT_RED = "FCE4D6"
WHITE_COLOR = "FFFFFF"

# Fonts
font_title = Font(name="Segoe UI", size=16, bold=True, color="1F497D")
font_meta_label = Font(name="Segoe UI", size=10, bold=True)
font_meta_val = Font(name="Segoe UI", size=10)
font_header = Font(name="Segoe UI", size=10, bold=True, color=WHITE)
font_lvl0 = Font(name="Segoe UI", size=10, bold=True)
font_lvl1 = Font(name="Segoe UI", size=10, bold=True)
font_lvl2 = Font(name="Segoe UI", size=10, italic=False)
font_lvl3 = Font(name="Segoe UI", size=9, italic=True, color="595959")
font_hierarchy = Font(name="Segoe UI", size=9, color="333333")

# Fills
fill_header = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
fill_lvl0 = PatternFill(start_color=LIGHT_BLUE_LVL0, end_color=LIGHT_BLUE_LVL0, fill_type="solid")
fill_lvl1 = PatternFill(start_color=LIGHT_BLUE_LVL1, end_color=LIGHT_BLUE_LVL1, fill_type="solid")
fill_lvl2 = PatternFill(start_color=LIGHT_GRAY_LVL2, end_color=LIGHT_GRAY_LVL2, fill_type="solid")
fill_lvl3 = PatternFill(start_color=WHITE_COLOR, end_color=WHITE_COLOR, fill_type="solid")

fill_green = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
fill_red = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")

# Alignments
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
align_hierarchy = Alignment(horizontal="left", vertical="center", wrap_text=True)

def align_with_indent(level):
    return Alignment(horizontal="left", vertical="center", indent=level * 2)

# Borders
thin_side = Side(border_style="thin", color="D3D3D3")
border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)


# ----------------- SHEET 1: RINGKASAN PAGU -----------------
ws_summary = wb.active
ws_summary.title = "Ringkasan Pagu"
ws_summary.views.sheetView[0].showGridLines = True

ws_summary["A1"] = "RINGKASAN PAGU ANGGARAN SATKER TAHUN 2026"
ws_summary["A1"].font = font_title
ws_summary.merge_cells("A1:J1")

# KPI Block (Total Pagu)
ws_summary["A3"] = "TOTAL PAGU SATKER"
ws_summary["A3"].font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
ws_summary["A3"].fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
ws_summary["A3"].alignment = align_center

total_satker_pagu = 0

headers_summary = [
    "No", "Kode Program", "Nama Program", "Kode Kegiatan", "Nama Kegiatan",
    "Kode KRO", "Nama KRO", "Kode Komponen", "Nama Komponen", "Pagu Komponen"
]

header_row = 6
for col_idx, h in enumerate(headers_summary, start=1):
    cell = ws_summary.cell(row=header_row, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_cell

row_idx = 7
num = 1
for prog in rka_data:
    prog_code = prog["text"].split("]")[0].replace("[", "").strip()
    prog_name = prog["text"].split("]")[1].strip()
    
    for keg in prog.get("kegiatans", []):
        for out in keg.get("outputs", []):
            for komp in out.get("komponens", []):
                ws_summary.cell(row=row_idx, column=1, value=num).alignment = align_center
                ws_summary.cell(row=row_idx, column=2, value=prog_code).alignment = align_center
                ws_summary.cell(row=row_idx, column=3, value=prog_name).alignment = align_left
                ws_summary.cell(row=row_idx, column=4, value=keg["code"]).alignment = align_center
                ws_summary.cell(row=row_idx, column=5, value=keg["name"]).alignment = align_left
                ws_summary.cell(row=row_idx, column=6, value=out["code"]).alignment = align_center
                ws_summary.cell(row=row_idx, column=7, value=out["name"]).alignment = align_left
                ws_summary.cell(row=row_idx, column=8, value=komp["code"]).alignment = align_center
                ws_summary.cell(row=row_idx, column=9, value=komp["name"]).alignment = align_left
                
                c_pagu = ws_summary.cell(row=row_idx, column=10, value=komp["pagu"])
                c_pagu.number_format = '#,##0'
                c_pagu.alignment = align_right
                
                for col_c in range(1, 11):
                    cell = ws_summary.cell(row=row_idx, column=col_c)
                    cell.border = border_cell
                    cell.font = Font(name="Segoe UI", size=10)
                
                total_satker_pagu += komp["pagu"]
                row_idx += 1
                num += 1

ws_summary.cell(row=row_idx, column=9, value="TOTAL").font = Font(name="Segoe UI", size=10, bold=True)
ws_summary.cell(row=row_idx, column=9).alignment = align_right
ws_summary.cell(row=row_idx, column=9).border = border_cell

c_tot = ws_summary.cell(row=row_idx, column=10, value=total_satker_pagu)
c_tot.font = Font(name="Segoe UI", size=10, bold=True)
c_tot.number_format = '#,##0'
c_tot.alignment = align_right
c_tot.border = border_cell

ws_summary["A4"] = total_satker_pagu
ws_summary["A4"].font = Font(name="Segoe UI", size=18, bold=True, color="1F497D")
ws_summary["A4"].number_format = 'Rp #,##0'
ws_summary["A4"].alignment = align_center
ws_summary.merge_cells("A3:C3")
ws_summary.merge_cells("A4:C4")

for col in ws_summary.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.row < 6:
            continue
        val_str = str(cell.value or '')
        if len(val_str) > max_len:
            max_len = len(val_str)
    ws_summary.column_dimensions[col_letter].width = max(max_len + 4, 10)

ws_summary.column_dimensions['C'].width = 30
ws_summary.column_dimensions['E'].width = 30
ws_summary.column_dimensions['G'].width = 30
ws_summary.column_dimensions['I'].width = 30


# ----------------- SHEET 2: SANDING RKA & RUP -----------------
ws_sanding = wb.create_sheet(title="Sanding RKA & RUP")
ws_sanding.views.sheetView[0].showGridLines = True

ws_sanding["A1"] = "PENYANDINGAN PAGU RKA DENGAN REALISASI RUP (TERUMUMKAN)"
ws_sanding["A1"].font = font_title
ws_sanding.merge_cells("A1:K1")

headers_sanding = [
    "No", "Program", "Kegiatan", "KRO (Output)", "Komponen", 
    "Key Otorisasi (MAK)", "Pagu RKA (A)", "Pagu RUP Terumumkan (B)", 
    "Selisih (A - B)", "Persentase (%)", "Status Evaluasi"
]

header_row = 4
for col_idx, h in enumerate(headers_sanding, start=1):
    cell = ws_sanding.cell(row=header_row, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_cell

row_idx = 5
num = 1
total_rup_pagu = 0

for prog in rka_data:
    prog_code = prog["text"].split("]")[0].replace("[", "").strip()
    
    for keg in prog.get("kegiatans", []):
        for out in keg.get("outputs", []):
            for komp in out.get("komponens", []):
                comp_key = f"{prog_code}.{keg['code']}.{out['code']}.{komp['code']}"
                
                # Sum RUP packets for this component
                rup_list = rup_by_comp.get(comp_key, [])
                rup_sum = sum(p["pagu"] for p in rup_list)
                total_rup_pagu += rup_sum
                
                selisih = komp["pagu"] - rup_sum
                pct = (rup_sum / komp["pagu"] * 100) if komp["pagu"] > 0 else 0
                
                ws_sanding.cell(row=row_idx, column=1, value=num).alignment = align_center
                ws_sanding.cell(row=row_idx, column=2, value=prog_code).alignment = align_center
                ws_sanding.cell(row=row_idx, column=3, value=keg["name"]).alignment = align_left
                ws_sanding.cell(row=row_idx, column=4, value=out["name"]).alignment = align_left
                ws_sanding.cell(row=row_idx, column=5, value=komp["name"]).alignment = align_left
                ws_sanding.cell(row=row_idx, column=6, value=comp_key).alignment = align_center
                
                c_pagu_rka = ws_sanding.cell(row=row_idx, column=7, value=komp["pagu"])
                c_pagu_rka.number_format = '#,##0'
                c_pagu_rka.alignment = align_right
                
                c_pagu_rup = ws_sanding.cell(row=row_idx, column=8, value=rup_sum)
                c_pagu_rup.number_format = '#,##0'
                c_pagu_rup.alignment = align_right
                
                c_sel = ws_sanding.cell(row=row_idx, column=9, value=selisih)
                c_sel.number_format = '#,##0'
                c_sel.alignment = align_right
                
                c_pct = ws_sanding.cell(row=row_idx, column=10, value=pct / 100)
                c_pct.number_format = '0.0%'
                c_pct.alignment = align_right
                
                status_cell = ws_sanding.cell(row=row_idx, column=11)
                status_cell.alignment = align_center
                
                if pct == 100:
                    status_cell.value = "Sesuai (100%)"
                    status_cell.fill = fill_green
                elif pct > 0:
                    status_cell.value = f"Parsial ({pct:.1f}%)"
                    status_cell.fill = fill_lvl1
                else:
                    status_cell.value = "Belum Diumumkan"
                    status_cell.fill = fill_red
                
                for col_c in range(1, 12):
                    cell = ws_sanding.cell(row=row_idx, column=col_c)
                    cell.border = border_cell
                    cell.font = Font(name="Segoe UI", size=9)
                
                row_idx += 1
                num += 1

# Total Row Sanding
ws_sanding.cell(row=row_idx, column=6, value="TOTAL").font = Font(name="Segoe UI", size=10, bold=True)
ws_sanding.cell(row=row_idx, column=6).alignment = align_right
ws_sanding.cell(row=row_idx, column=6).border = border_cell

for col_c in (7, 8, 9):
    val = total_satker_pagu if col_c == 7 else (total_rup_pagu if col_c == 8 else (total_satker_pagu - total_rup_pagu))
    cell = ws_sanding.cell(row=row_idx, column=col_c, value=val)
    cell.font = Font(name="Segoe UI", size=10, bold=True)
    cell.number_format = '#,##0'
    cell.alignment = align_right
    cell.border = border_cell

tot_pct = (total_rup_pagu / total_satker_pagu) if total_satker_pagu > 0 else 0
c_tot_pct = ws_sanding.cell(row=row_idx, column=10, value=tot_pct)
c_tot_pct.font = Font(name="Segoe UI", size=10, bold=True)
c_tot_pct.number_format = '0.0%'
c_tot_pct.alignment = align_right
c_tot_pct.border = border_cell

# Card Summary Info
ws_sanding["A2"] = f"Total Pagu RKA: Rp {total_satker_pagu:,} | Terumumkan di RUP: Rp {total_rup_pagu:,} ({tot_pct*100:.1f}%)"
ws_sanding["A2"].font = Font(name="Segoe UI", size=10, italic=True)
ws_sanding.merge_cells("A2:K2")

ws_sanding.column_dimensions['A'].width = 6
ws_sanding.column_dimensions['B'].width = 10
ws_sanding.column_dimensions['C'].width = 25
ws_sanding.column_dimensions['D'].width = 25
ws_sanding.column_dimensions['E'].width = 25
ws_sanding.column_dimensions['F'].width = 18
ws_sanding.column_dimensions['G'].width = 15
ws_sanding.column_dimensions['H'].width = 22
ws_sanding.column_dimensions['I'].width = 15
ws_sanding.column_dimensions['J'].width = 12
ws_sanding.column_dimensions['K'].width = 18


# ----------------- SHEET 3: DAFTAR PAKET RUP -----------------
ws_rup = wb.create_sheet(title="Daftar Paket RUP")
ws_rup.views.sheetView[0].showGridLines = True

ws_rup["A1"] = "DAFTAR PAKET PENYEDIA EXISTING YANG TERUMUMKAN (RUP)"
ws_rup["A1"].font = font_title
ws_rup.merge_cells("A1:L1")

headers_rup = [
    "No", "ID Paket", "Nama Kegiatan (RUP)", "Nama Paket", "Pagu RUP (Rp)",
    "Waktu Pemilihan", "Sumber Dana", "A (Draft PPK)", "FD (Final Draft PPK)",
    "U (Terumumkan KPA)", "Kode Otorisasi (MAK)", "Mata Anggaran Mapped"
]

header_row = 4
for col_idx, h in enumerate(headers_rup, start=1):
    cell = ws_rup.cell(row=header_row, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_cell

row_idx = 5
for idx, p in enumerate(rup_packets, start=1):
    ws_rup.cell(row=row_idx, column=1, value=idx).alignment = align_center
    ws_rup.cell(row=row_idx, column=2, value=p["id"]).alignment = align_center
    ws_rup.cell(row=row_idx, column=3, value=p["keg_name"]).alignment = align_left
    ws_rup.cell(row=row_idx, column=4, value=p["name"]).alignment = align_left
    
    c_pagu = ws_rup.cell(row=row_idx, column=5, value=p["pagu"])
    c_pagu.number_format = '#,##0'
    c_pagu.alignment = align_right
    
    ws_rup.cell(row=row_idx, column=6, value=p["waktu"]).alignment = align_center
    ws_rup.cell(row=row_idx, column=7, value=p["sumber_dana"]).alignment = align_center
    
    # Status Indicators A, FD, U
    ws_rup.cell(row=row_idx, column=8, value="✓" if p["aktif"] else "").alignment = align_center
    ws_rup.cell(row=row_idx, column=9, value="✓" if p["fd"] else "").alignment = align_center
    ws_rup.cell(row=row_idx, column=10, value="✓" if p["umumkan"] else "").alignment = align_center
    
    ws_rup.cell(row=row_idx, column=11, value=p["mak"]).alignment = align_center
    ws_rup.cell(row=row_idx, column=12, value=p["comp_key"]).alignment = align_center
    
    for col_c in range(1, 13):
        cell = ws_rup.cell(row=row_idx, column=col_c)
        cell.border = border_cell
        cell.font = Font(name="Segoe UI", size=9)
        
    row_idx += 1

# Total Row for RUP
ws_rup.cell(row=row_idx, column=4, value="TOTAL").font = Font(name="Segoe UI", size=10, bold=True)
ws_rup.cell(row=row_idx, column=4).alignment = align_right
ws_rup.cell(row=row_idx, column=4).border = border_cell

c_tot_rup = ws_rup.cell(row=row_idx, column=5, value=total_rup_pagu)
c_tot_rup.font = Font(name="Segoe UI", size=10, bold=True)
c_tot_rup.number_format = '#,##0'
c_tot_rup.alignment = align_right
c_tot_rup.border = border_cell

ws_rup.column_dimensions['A'].width = 5
ws_rup.column_dimensions['B'].width = 12
ws_rup.column_dimensions['C'].width = 25
ws_rup.column_dimensions['D'].width = 35
ws_rup.column_dimensions['E'].width = 15
ws_rup.column_dimensions['F'].width = 15
ws_rup.column_dimensions['G'].width = 12
ws_rup.column_dimensions['H'].width = 15
ws_rup.column_dimensions['I'].width = 18
ws_rup.column_dimensions['J'].width = 18
ws_rup.column_dimensions['K'].width = 30
ws_rup.column_dimensions['L'].width = 20


# ----------------- SHEET 4 & 5: DETAIL PER PROGRAM -----------------
for prog in rka_data:
    prog_code = prog["text"].split("]")[0].replace("[", "").strip()
    sheet_name = f"Detail - {prog_code}"
    ws = wb.create_sheet(title=sheet_name)
    ws.views.sheetView[0].showGridLines = True
    
    # Title
    ws["A1"] = f"DETAIL RENCANA KERJA ANGGARAN (RKA) - PROGRAM {prog_code}"
    ws["A1"].font = font_title
    ws.merge_cells("A1:M1")
    
    ws["A3"] = "Program:"
    ws["A3"].font = font_meta_label
    ws["B3"] = prog["text"]
    ws["B3"].font = font_meta_val
    ws.merge_cells("B3:M3")
    
    headers_det = [
        "Kegiatan", "KRO (Output)", "Komponen", "Kode (P/K/O/SO/K/SK/A/D)",
        "Uraian", "Uraian Sebelum Revisi", "Pagu", "Pagu Sebelum Revisi",
        "P", "S", "Multiyears", "NP", "Gaji"
    ]
    
    # Write Headers
    header_row = 5
    for col_idx, h in enumerate(headers_det, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_cell
        
    curr_row = 6
    
    for keg in prog.get("kegiatans", []):
        keg_text = f"[{keg['code']}] {keg['name']}"
        for out in keg.get("outputs", []):
            kro_text = f"[{out['code']}] {out['name']}"
            for komp in out.get("komponens", []):
                komp_text = f"[{komp['code']}] {komp['name']}"
                
                start_row = curr_row
                
                for row_data in komp.get("rows", []):
                    level = row_data["level"]
                    code = row_data["code"]
                    desc = row_data["desc"]
                    desc_prev = row_data["descPrev"]
                    pagu = row_data["pagu"]
                    pagu_prev = row_data["paguPrev"]
                    
                    p_ch = row_data["p_ch"]
                    s_ch = row_data["s_ch"]
                    my_ch = row_data["my_ch"]
                    np_ch = row_data["np_ch"]
                    gj_ch = row_data["gj_ch"]
                    
                    ws.cell(row=curr_row, column=4, value=code)
                    ws.cell(row=curr_row, column=5, value=desc)
                    ws.cell(row=curr_row, column=6, value=desc_prev)
                    
                    if isinstance(pagu, (int, float)):
                        ws.cell(row=curr_row, column=7, value=pagu).number_format = '#,##0'
                    else:
                        ws.cell(row=curr_row, column=7, value=pagu)
                        
                    if isinstance(pagu_prev, (int, float)):
                        ws.cell(row=curr_row, column=8, value=pagu_prev).number_format = '#,##0'
                    else:
                        ws.cell(row=curr_row, column=8, value=pagu_prev)
                        
                    ws.cell(row=curr_row, column=9, value="✓" if p_ch else "")
                    ws.cell(row=curr_row, column=10, value="✓" if s_ch else "")
                    ws.cell(row=curr_row, column=11, value="✓" if my_ch else "")
                    ws.cell(row=curr_row, column=12, value="✓" if np_ch else "")
                    ws.cell(row=curr_row, column=13, value="✓" if gj_ch else "")
                    
                    if level == 0:
                        row_fill = fill_lvl0
                        row_font = font_lvl0
                    elif level == 1:
                        row_fill = fill_lvl1
                        row_font = font_lvl1
                    elif level == 2:
                        row_fill = fill_lvl2
                        row_font = font_lvl2
                    else:
                        row_fill = fill_lvl3
                        row_font = font_lvl3
                        
                    for col_idx in range(4, 14):
                        cell = ws.cell(row=curr_row, column=col_idx)
                        cell.fill = row_fill
                        cell.font = row_font
                        cell.border = border_cell
                        
                        if col_idx == 4:
                            cell.alignment = align_center
                        elif col_idx == 5:
                            cell.alignment = align_with_indent(level)
                        elif col_idx == 6:
                            cell.alignment = align_with_indent(level)
                        elif col_idx in (7, 8):
                            cell.alignment = align_right
                        else:
                            cell.alignment = align_center
                    
                    curr_row += 1
                
                end_row = curr_row - 1
                
                if end_row >= start_row:
                    ws.cell(row=start_row, column=1, value=keg_text)
                    ws.cell(row=start_row, column=2, value=kro_text)
                    ws.cell(row=start_row, column=3, value=komp_text)
                    
                    for r in range(start_row, end_row + 1):
                        for c in (1, 2, 3):
                            cell = ws.cell(row=r, column=c)
                            cell.font = font_hierarchy
                            cell.alignment = align_hierarchy
                            cell.border = border_cell
                            cell.fill = PatternFill(start_color="FAFBFD", end_color="FAFBFD", fill_type="solid")
                    
                    if end_row > start_row:
                        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                        ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
                        ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)

    # Dimensions
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 45
    ws.column_dimensions['F'].width = 45
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    for col_c in range(9, 14):
        col_letter = get_column_letter(col_c)
        ws.column_dimensions[col_letter].width = 8

try:
    wb.save("rekap_rka_lengkap.xlsx")
    print("Consolidated Excel with RUP data updated successfully!")
except PermissionError:
    import time
    alt_name = f"rekap_rka_lengkap_{int(time.time())}.xlsx"
    wb.save(alt_name)
    print(f"Permission denied on original file. Saved as alternative: {alt_name}")
